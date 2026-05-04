

"""
============================================================
SBERT-Powered Call Coverage Checker (Mandarin Analysis)
============================================================

This script is an advanced version of the call coverage checker, specifically
optimized for Mandarin (Simplified Chinese) transcripts. It replaces the traditional
TF-IDF algorithm with a powerful Sentence-BERT (SBERT) model for superior
semantic similarity understanding.

Key Upgrades:
1. **Core Engine: Sentence-BERT (SBERT):** Uses 'paraphrase-multilingual-MiniLM-L12-v2'
   for deep semantic understanding, moving from keyword matching to intent matching.
2. **Mandarin Tokenizer: Jieba:** Replaces the Cantonese-specific 'pycantonese'
   with 'jieba', the industry standard for Mandarin word segmentation.
3. **Streamlined & Focused:** All TF-IDF related logic has been removed, resulting
   in a cleaner and more modern codebase.
4. **Enhanced Pattern Recognition:** Supports date/numeric pattern recognition for scoring boosts.
5. **SBERT优化的简化验证策略:** 采用"简化的单向验证"而非复杂的双重打分：
   - 信任SBERT的语义理解能力，以"单句匹配"为主
   - 通过放宽上下文窗口(350字符)为SBERT提供丰富上下文
   - 让对话组本身扮演"整段"角色，无需复杂的双重验证
6. **3-Pass Speaker Grouping:** 优化的3-pass分组策略，为SBERT提供最佳语义环境。
7. **Configurable Input Files:** Supports separate call text and script files.

============================================================
"""

# =============================
# Imports & Dependencies
# =============================
import pandas as pd
import jieba  # Replaced pycantonese with jieba for Mandarin
import re
import os
import sys
import unicodedata
from difflib import SequenceMatcher
import numpy as np
from collections import Counter  # 移到顶部避免循环内重复import
import tracemalloc
import time

# =============================
# Project Path Setup
# =============================
# Project root = folder containing this file (flat layout: all .py in one directory)
project_root = os.path.dirname(os.path.abspath(__file__))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

# Import project modules
import dictionaries
from config import (
    SBERT_MODEL_PATH,
    SCRIPT_EMBEDDINGS_PATH,
    COMPOSITE_POINT_KEY_SEPARATOR,
    COMPOSITE_POINT_NOTE_COLUMN,
    should_use_composite_point_key,
)

# =============================
# Project Root Path Setup
# =============================
# Project root and dictionaries already set up at top of file

# --- SBERT related imports ---
# Ensure 'sentence-transformers' is installed in your environment
try:
    from sentence_transformers import SentenceTransformer, util
except ImportError:
    print("❌ Error: 'sentence-transformers' package not found.")
    print("Please install it in your environment using the offline method.")
    sys.exit(1)

# Ensure proper encoding for Windows
if sys.platform.startswith('win'):
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.detach())
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.detach())

# Dictionaries imported at top of file

# =============================
# USER CONFIGURATION SECTION
# =============================
"""
CONFIGURE YOUR DISCUSSION POINTS FOR ENHANCED SCORING HERE:
============================================================

Instructions:
1. Add discussion point names that you want to enhance with date/numeric pattern recognition
2. Set 'date_boost': score boost (0.05-0.25) for points that should get bonus for date patterns
3. Set 'numeric_boost': score boost (0.05-0.25) for points that should get bonus for numeric patterns
4. Set both to 0.0 if you don't want that type of enhancement

Examples:
- For points about dates/deadlines: set date_boost > 0, numeric_boost = 0
- For points about prices/amounts: set numeric_boost > 0, date_boost = 0
- For mixed points: set both > 0
"""

# MODIFY THIS DICTIONARY TO CONFIGURE YOUR DISCUSSION POINTS:
USER_ENHANCED_DISCUSSION_POINTS = {
    # Example entries - modify these as needed for Mandarin:
    '价格确认': {'date_boost': 0.0, 'numeric_boost': 0.20},
    '风险评估': {'date_boost': 0.15, 'numeric_boost': 0.0},
    '交易价格确认': {'date_boost': 0.0, 'numeric_boost': 0.18},
    '交易数量确认': {'date_boost': 0.0, 'numeric_boost': 0.20},
    '产品信息确认': {'date_boost': 0.0, 'numeric_boost': 0.12},
    '重要事项声明': {'date_boost': 0.12, 'numeric_boost': 0.0},
    '财务状况评估': {'date_boost': 0.0, 'numeric_boost': 0.15},
    
    # Template for adding more:
    # 'Your Discussion Point Name': {'date_boost': 0.1, 'numeric_boost': 0.15},
}

# =============================
# INPUT FILE CONFIGURATION
# =============================
"""
CONFIGURE YOUR INPUT FILE PATHS HERE:
====================================

Instructions:
1. Set CALL_TEXT_FILE_PATH to the path of your call transcript file
2. Set SCRIPT_FILE_PATH to the path of your script/required points file
3. Use forward slashes (/) or raw strings for Windows paths
4. Can be relative paths (from script location) or absolute paths

File Format Requirements:
- Call text file: Must contain columns 'Speaker Roles', 'Transcription', 'Segment Start Time', 'Segment End Time'
- Script file: Must contain columns 'Required_Discussion_Point', 'Standard_Script'
"""

# ==========================================================
# CONFIGURATION (Paths are now passed via function parameters)
# ==========================================================
# Note: File paths and sheet names are now provided via run_analysis() function parameters
# when called from run_batch_analysis.py or other batch processing systems.

# =============================
# SBERT MODEL CONFIGURATION
# =============================
"""
SBERT MODEL PATH CONFIGURATION:
===============================

Set the path to your local SBERT model folder.
The model should be the 'paraphrase-multilingual-MiniLM-L12-v2' model.
"""

# SBERT paths come directly from config.py (fail-fast for profile consistency)
_ = SBERT_MODEL_PATH, SCRIPT_EMBEDDINGS_PATH

# =============================
# UTILITY FUNCTIONS
# =============================
def print_checkpoint(step_num, description, start_time=None):
    """Progress tracking function"""
    current_time = time.time()
    timestamp = time.strftime('%H:%M:%S')
    if start_time is not None:
        elapsed = current_time - start_time
        print(f"✅ CHECKPOINT {step_num}: {description} (Completed at: {timestamp}, Duration: {elapsed:.2f}s)")
    else:
        print(f"🚀 CHECKPOINT {step_num}: {description} (Started at: {timestamp})")
    return current_time

# ==========================================================
# Class: SbertCallCoverageChecker
# ==========================================================
class SbertCallCoverageChecker:
    def __init__(self, model_path):
        """
        Initialization: Loads dictionaries and the SBERT model.
        """
        # Load dictionaries (from shared dictionaries.py imported at top)
        self.mandarin_synonyms = dictionaries.cantonese_synonyms  # Use shared synonyms
        self.important_keywords = dictionaries.important_keywords
        # Language-specific stopwords will be loaded in load_call_specific_weights
        self.stopwords = None  # Will be set based on detected language
        
        # Multi-product weight management
        self.current_language = None
        self.current_product = None
        self.current_weights = {}  # Will be loaded by load_call_specific_weights
        
        # Performance optimization caches
        self._preprocessed_texts_cache = {}
        self._similarity_cache = {}
        self._reverse_synonym_map = None
        
        # SBERT embeddings cache for precomputed script vectors
        self.script_embeddings = None
        
        # 改动点1: 新增轻量缓存 (最小侵入式)
        self._preprocess_cache = {}
        self._expand_keywords_cache = {}
        self._jieba_tokens_cache = {}
        self._stopwords_MAN = None  # 懒加载

        # --- Load the SBERT model ---
        if not os.path.exists(model_path):
            print(f"❌ FATAL ERROR: SBERT model path does not exist: '{model_path}'")
            print("Please update the SBERT_MODEL_PATH variable in the script.")
            sys.exit(1)
        
        try:
            print(f"Loading SBERT model from: {model_path} ... (This may take a moment)")
            self.sbert_model = SentenceTransformer(model_path)
            print("✅ SBERT model loaded successfully.")
        except Exception as e:
            print(f"❌ FATAL ERROR: Failed to load SBERT model from '{model_path}'.")
            print(f"Error details: {e}")
            print("Please ensure the model folder is complete and all dependencies are installed.")
            sys.exit(1)

    # ------------------------------------------------------
    # Text Preprocessing & Tokenization (Adapted for Mandarin)
    # ------------------------------------------------------
    def preprocess_text(self, text, mode='comparison', text_type='call'):
        """
        统一文本预处理函数，支持多种模式（普通话版本）
        
        Args:
            text: 输入文本
            mode: 处理模式
                - 'comparison': 用于比较分析（去除英文，script和call统一处理）
                - 'display': 用于显示（保留英文）
            text_type: 文本类型 ('call' 或 'script')，主要用于缓存区分
                
        Returns:
            处理后的文本
        """
        text_str = str(text)
        # 创建缓存键，包含模式和文本类型
        cache_key = f"{mode}_{text_type}_{text_str}"
        
        # 检查缓存
        if cache_key in self._preprocessed_texts_cache:
            return self._preprocessed_texts_cache[cache_key]
        
        # ENG分支：调用english_resource进行英文预处理
        if hasattr(self, 'current_language') and self.current_language == 'ENG':
            try:
                # 对于英文，使用轻量归一化，不剔除英文字母
                processed = unicodedata.normalize('NFKC', text_str)
                # 基础标点标准化
                processed = re.sub(r'[,，]+', ',', processed)  # 标准化逗号
                processed = re.sub(r'[.。]+', '.', processed)  # 标准化句号
                print(f"🔤 ENG preprocess: '{text_str[:50]}...' -> '{processed[:50]}...'")
            except Exception as e:
                # 异常处理：进行基础英文处理
                print(f"⚠️  ENG preprocess error: {e}, using basic processing")
                processed = unicodedata.normalize('NFKC', text_str)
                processed = re.sub(r'[,，]+', ',', processed)
                processed = re.sub(r'[.。]+', '.', processed)
        else:
            # MAN/CAN分支：保持原有逻辑不变
            if mode == 'comparison':
                # 比较模式：去除英文字母（script和call统一处理，确保比较一致性）
                processed = re.sub(r'[^\u4e00-\u9fff,.。，。%()（）]+', '', text_str)
            elif mode == 'display':
                # 显示模式：保留英文字母（用于sentence level analysis等显示用途）
                processed = re.sub(r'[^\u4e00-\u9fffa-zA-Z,.。，。%()（）]+', '', text_str)
            else:
                # 默认使用比较模式
                processed = re.sub(r'[^\u4e00-\u9fff,.。，。%()（）]+', '', text_str)
            
            # 普通话版本不需要error_patterns处理
            
            # 普通话特定的标准化
            processed = re.sub(r'[,，]+', '，', processed)  # 标准化逗号
            processed = re.sub(r'[.。]+', '。', processed)  # 标准化句号
        
        result = processed.strip()
        
        # 缓存结果（限制缓存大小以防内存溢出）
        if len(self._preprocessed_texts_cache) < 10000:
            self._preprocessed_texts_cache[cache_key] = result
        
        return result

    def _build_reverse_synonym_map(self):
        """
        Build reverse synonym mapping for O(1) lookup performance.
        Maps each synonym to its canonical key and all related synonyms.
        """
        if self._reverse_synonym_map is not None:
            return self._reverse_synonym_map
            
        reverse_map = {}
        for key, synonyms in self.mandarin_synonyms.items():
            # Map the key to itself and all synonyms
            all_variants = {key} | set(synonyms)
            reverse_map[key] = all_variants
            
            # Map each synonym to the same set
            for synonym in synonyms:
                reverse_map[synonym] = all_variants
                
        self._reverse_synonym_map = reverse_map
        return reverse_map
    
    def expand_keywords(self, text):
        """
        Expand text with language-specific synonyms for robust matching.
        - For ENG: Uses english_resource expansion strategies
        - For MAN/CAN: Uses jieba and pre-computed reverse mapping
        """
        # ENG分支：使用english_resource扩展，不使用否定bigram
        if hasattr(self, 'current_language') and self.current_language == 'ENG':
            try:
                import english_resource
                # 使用english_resource进行分词和同义词扩展，不使用否定bigram
                tokens = english_resource.normalize_and_tokenize_en(text)
                expanded_tokens = english_resource.expand_phrases_and_synonyms_en(tokens)
                # 用dictionaries.get_stopwords('ENG')再次过滤
                eng_stopwords = dictionaries.get_stopwords('ENG')
                filtered_expanded = {token for token in expanded_tokens if token not in eng_stopwords}
                print(f"🔤 ENG expand: '{text[:30]}...' -> {len(filtered_expanded)} expanded tokens: {list(filtered_expanded)[:5]}...")
                return filtered_expanded
            except ImportError:
                print(f"⚠️  english_resource not available, using basic English expansion")
                # 基础英文处理兜底
                import re
                tokens = re.findall(r'\w+', text.lower())
                eng_stopwords = dictionaries.get_stopwords('ENG')
                return set(token for token in tokens if token not in eng_stopwords and len(token) > 1)
            except Exception as e:
                print(f"⚠️  ENG expand error: {e}, using basic expansion")
                import re
                tokens = re.findall(r'\w+', text.lower())
                eng_stopwords = dictionaries.get_stopwords('ENG')
                return set(token for token in tokens if token not in eng_stopwords and len(token) > 1)
        else:
            # MAN/CAN分支：保持原有逻辑不变
            reverse_map = self._build_reverse_synonym_map()
            expanded_tokens = set()
            tokens = jieba.lcut(text)  # Using jieba for Mandarin
            
            for token in tokens:
                expanded_tokens.add(token)
                # Use reverse mapping for fast lookup
                if token in reverse_map:
                    expanded_tokens.update(reverse_map[token])
            
            return expanded_tokens

    # 改动点1: 薄包装函数 (最小侵入式)
    def _preprocess_memo(self, text):
        """预处理缓存版本，避免重复预处理"""
        if text in self._preprocess_cache:
            return self._preprocess_cache[text]
        
        result = self.preprocess_text(text)
        # 简单的缓存大小控制
        if len(self._preprocess_cache) < 5000:
            self._preprocess_cache[text] = result
        return result

    def normalize_script_embedding_keys(self):
        """
        Normalize precomputed script embedding keys to comparison-mode clean text.
        This makes embedding cache lookup robust against irrelevant raw-text differences.
        """
        if not self.script_embeddings:
            self.script_embeddings = {}
            return 0, 0, 0

        if not isinstance(self.script_embeddings, dict):
            print("⚠️  Invalid script embeddings format, falling back to empty dict")
            self.script_embeddings = {}
            return 0, 0, 0

        original_count = len(self.script_embeddings)
        normalized = {}
        collisions = 0

        for raw_key, embedding in self.script_embeddings.items():
            clean_key = self.preprocess_text(raw_key, mode='comparison', text_type='script')
            if not clean_key:
                continue
            if clean_key in normalized:
                collisions += 1
                continue
            normalized[clean_key] = embedding

        self.script_embeddings = normalized
        return original_count, len(normalized), collisions
    
    def _expand_keywords_memo(self, clean_text):
        """关键词扩展缓存版本，避免重复扩展"""
        if clean_text in self._expand_keywords_cache:
            return self._expand_keywords_cache[clean_text]
        
        result = self.expand_keywords(clean_text)
        # 转为frozenset以支持缓存
        frozen_result = frozenset(result)
        if len(self._expand_keywords_cache) < 5000:
            self._expand_keywords_cache[clean_text] = frozen_result
        return frozen_result
    
    def _jieba_tokens_memo(self, text):
        """jieba分词缓存版本，避免重复分词，并懒加载停用词"""
        if text in self._jieba_tokens_cache:
            return self._jieba_tokens_cache[text]
        
        # 懒加载停用词
        if self._stopwords_MAN is None:
            self._stopwords_MAN = set(dictionaries.get_stopwords('MAN'))
        
        tokens = jieba.lcut(text)
        # 过滤停用词
        filtered_tokens = [token for token in tokens if token not in self._stopwords_MAN and len(token) > 1]
        
        if len(self._jieba_tokens_cache) < 5000:
            self._jieba_tokens_cache[text] = filtered_tokens
        return filtered_tokens
    
    def tokenize_text(self, text):
        """
        Tokenize text using jieba for Mandarin or english_resource for English.
        """
        # ENG分支：使用english_resource + dictionaries.get_stopwords统一过滤
        if hasattr(self, 'current_language') and self.current_language == 'ENG':
            try:
                import english_resource
                # 使用english_resource进行英文分词，然后用统一的停用词过滤
                tokens = english_resource.normalize_and_tokenize_en(text)
                # 再次用dictionaries.get_stopwords('ENG')过滤，确保一致性
                eng_stopwords = dictionaries.get_stopwords('ENG')
                filtered_tokens = [token for token in tokens if token not in eng_stopwords and len(token) > 1]
                print(f"🔤 ENG tokenize: '{text[:30]}...' -> {len(filtered_tokens)} tokens: {filtered_tokens[:5]}...")
                return filtered_tokens
            except ImportError:
                print(f"⚠️  english_resource not available, using basic English tokenization")
                # 基础英文分词兜底
                import re
                tokens = re.findall(r'\w+', text.lower())
                eng_stopwords = dictionaries.get_stopwords('ENG')
                return [token for token in tokens if token not in eng_stopwords and len(token) > 1]
            except Exception as e:
                print(f"⚠️  ENG tokenize error: {e}, using basic tokenization")
                import re
                tokens = re.findall(r'\w+', text.lower())
                eng_stopwords = dictionaries.get_stopwords('ENG')
                return [token for token in tokens if token not in eng_stopwords and len(token) > 1]
        else:
            # MAN/CAN分支：保持原有逻辑不变
            tokens = jieba.lcut(text)  # Using jieba for Mandarin
            # Filter out stopwords and very short tokens
            business_tokens = [token for token in tokens if token not in self.stopwords and len(token) > 1]
            return business_tokens

    # ------------------------------------------------------
    # Multi-Product Weight Management
    # ------------------------------------------------------
    def detect_language_from_filename(self, file_path):
        """
        Detect language from CSV filename (Mandarin version only supports _M and _E)
        Examples:
        - "xxxxxx_M.wav.csv" -> "MAN" (Mandarin)
        - "xxxxxx_E.wav.csv" -> "ENG" (English)
        - "xxxxxx_C.wav.csv" -> Rejected (Cantonese not supported in Mandarin version)
        """
        filename = os.path.basename(file_path)
        # Remove .csv extension
        name_without_csv = filename.replace('.csv', '')
        # Remove .wav if present
        name_without_wav = name_without_csv.replace('.wav', '')
        
        # Get the last character after underscore
        if name_without_wav.endswith('_M'):
            return "MAN"
        elif name_without_wav.endswith('_E'):
            return "ENG"
        elif name_without_wav.endswith('_C'):
            print(f"❌ Error: Cantonese files (_C) are not supported in the Mandarin version.")
            print(f"Please use the Cantonese version (improved_call_coverage_checker.py) for _C files.")
            return None
        else:
            return None

    def detect_product_type_from_script(self, script_df, script_sheet_name=None):
        """
        Detect product type from script DataFrame based on script_sheet_name.
        Extracts product name from script_sheet_name (before underscore)
        Example: 'SID CPI3_MAN' -> product = 'SID CPI3', language = 'MAN'
        """
        if script_sheet_name and '_' in script_sheet_name:
            # Extract product name (before underscore)
            product_name = script_sheet_name.split('_')[0].strip()
            print(f"🔍 Detected product from sheet name '{script_sheet_name}': '{product_name}'")
            return product_name
        elif script_sheet_name:
            print(f"⚠️  Sheet name '{script_sheet_name}' doesn't contain underscore, using as-is: {script_sheet_name}")
            return script_sheet_name
        else:
            print(f"⚠️  No script_sheet_name provided, cannot detect product type")
            return "Unknown"

    def load_call_specific_weights(self, call_file_path, script_df, script_sheet_name=None):
        """
        Load specific weights and language-specific configurations based on call file and script
        """
        print("🔍 Loading call-specific weights and language configurations...")
        
        # Use shared dictionary (imported at top)
        
        # 1. Detect language from filename
        self.current_language = self.detect_language_from_filename(call_file_path)
        
        # 1b. If language detection from filename fails, try to detect from sheet name
        if not self.current_language and script_sheet_name:
            try:
                if '_' in script_sheet_name:
                    # Extract language code (after underscore)
                    language_code = script_sheet_name.split('_')[1].strip()
                    language_mapping = {'MAN': 'MAN', 'ENG': 'ENG'}  # Mandarin version only supports MAN and ENG
                    if language_code in language_mapping:
                        self.current_language = language_mapping[language_code]
                        print(f"🔍 Detected language from sheet name '{script_sheet_name}': '{self.current_language}'")
                    elif language_code == 'CAN':
                        print(f"❌ Error: Cantonese files (_CAN) are not supported in the Mandarin version.")
                        print(f"Please use the Cantonese version (improved_call_coverage_checker.py) for _CAN files.")
                        return None
            except Exception as e:
                print(f"⚠️  Error detecting language from sheet name: {e}")
        
        # 2. Load language-specific stopwords
        if self.current_language:
            try:
                # 统一使用dictionaries.get_stopwords，包括ENG
                self.stopwords = dictionaries.get_stopwords(self.current_language)
                print(f"✅ Loaded {self.current_language} stopwords: {len(self.stopwords)} words")
            except Exception as e:
                print(f"⚠️  Error loading {self.current_language} stopwords: {e}")
                print(f"⚠️  Using default Mandarin stopwords")
                self.stopwords = dictionaries.get_stopwords('MAN')
        else:
            # Default to Mandarin for this version
            self.stopwords = dictionaries.get_stopwords('MAN')
        
        # 3. Detect product type from script
        self.current_product = self.detect_product_type_from_script(script_df, script_sheet_name)
        
        # 4. Load corresponding weights
        # Trust get_product_weights to handle all fallback logic internally
        if self.current_language and self.current_product:
                self.current_weights = dictionaries.get_product_weights(
                    self.current_language, self.current_product
                )
                print(f"✅ Loaded weights for {self.current_language}:{self.current_product}")
                print(f"📊 Using {len(self.current_weights)} term weights")
        else:
            # Only fall back to general weights if language/product detection completely fails
            if not self.current_language:
                print(f"⚠️  Could not detect language from filename: {call_file_path}")
                if script_sheet_name:
                    print(f"⚠️  Also could not detect language from sheet name: {script_sheet_name}")
                else:
                    print(f"⚠️  No script sheet name provided for language detection")
            if not self.current_product:
                if script_sheet_name:
                    print(f"⚠️  Could not detect product type from script sheet name: {script_sheet_name}")
                else:
                    print(f"⚠️  No script sheet name provided for product detection")
            print(f"⚠️  Using general fallback weights")
            self.current_weights = getattr(dictionaries, 'term_importance', {})

    # ------------------------------------------------------
    # Salescall System Audio Recording Detection
    # ------------------------------------------------------
    def detect_system_audio_recordings(self, call_df):
        """
        Salescall Detection: Identify system audio recordings in sales calls.
        
        Logic:
        - Excludes speaker_1 and speaker_2 (main participants)
        - For remaining speakers, finds consecutive segments with:
          * 10+ consecutive rows by same speaker
          * Average sentence length >= 20 characters
        - If any segment meets criteria, marks ALL rows for that speaker as recordings
          (handles interruptions like coughing between recording segments)
        
        Args:
            call_df: DataFrame with call data
            
        Returns:
            pandas.Series: Boolean series indicating system audio recording rows
        """
        print("🔍 Salescall: Starting system audio recording detection...")
        
        speaker_col = 'Speaker Roles'
        text_col = 'Transcription'
        
        if speaker_col not in call_df.columns or text_col not in call_df.columns:
            print("⚠️  Salescall: Required columns not found, skipping recording detection")
            return pd.Series([False] * len(call_df), index=call_df.index)
        
        # Initialize recording detection array
        is_recording = pd.Series([False] * len(call_df), index=call_df.index)
        
        # Get all speakers except main participants (SPEAKER_1 and SPEAKER_2 are sales and customer)
        # Normalize speaker names to uppercase for consistent comparison
        all_speakers = call_df[speaker_col].dropna().unique()
        normalized_speakers = [str(s).upper() for s in all_speakers]
        main_speakers = ['SPEAKER_1', 'SPEAKER_2']
        system_speakers = [s for s in all_speakers if str(s).upper() not in main_speakers]
        
        print(f"📊 Salescall: Found {len(system_speakers)} potential system speakers: {system_speakers}")
        
        # Track speakers identified as system recordings
        recording_speakers = set()
        
        for speaker in system_speakers:
            speaker_rows = call_df[call_df[speaker_col] == speaker]
            if len(speaker_rows) < 10:  # Need at least 10 rows
                continue
            
            # Find consecutive segments for this speaker
            speaker_indices = speaker_rows.index.tolist()
            consecutive_segments = self._find_consecutive_segments(speaker_indices)
            
            # Check if any segment meets recording criteria
            speaker_is_recording = False
            for segment_indices in consecutive_segments:
                if len(segment_indices) >= 10:  # At least 10 consecutive rows
                    # Check average sentence length
                    segment_texts = call_df.loc[segment_indices, text_col].fillna('')
                    text_lengths = [len(str(text).strip()) for text in segment_texts]
                    avg_length = sum(text_lengths) / len(text_lengths) if text_lengths else 0
                    
                    if avg_length >= 20:  # Average length >= 20 characters
                        speaker_is_recording = True
                        print(f"🎵 Salescall: Detected recording segment - Speaker: {speaker}, "
                              f"Segment rows: {len(segment_indices)}, Avg length: {avg_length:.1f}")
                        break  # Found qualifying segment, no need to check others
            
            # If speaker qualifies, mark ALL their rows as recordings
            if speaker_is_recording:
                recording_speakers.add(speaker)
                all_speaker_indices = speaker_rows.index.tolist()
                is_recording.loc[all_speaker_indices] = True
                print(f"📢 Salescall: Marked ALL {len(all_speaker_indices)} rows for speaker '{speaker}' as recordings")
        
        total_recording_rows = is_recording.sum()
        print(f"✅ Salescall: Detection complete - {total_recording_rows} rows marked as system recordings")
        print(f"📊 Salescall: Recording speakers identified: {list(recording_speakers)}")
        
        return is_recording
    
    def _find_consecutive_segments(self, indices):
        """
        Salescall Helper: Find consecutive segments in a list of indices.
        
        Args:
            indices: List of row indices
            
        Returns:
            List of lists, each containing consecutive indices
        """
        if not indices:
            return []
        
        segments = []
        current_segment = [indices[0]]
        
        for i in range(1, len(indices)):
            if indices[i] == indices[i-1] + 1:  # Consecutive
                current_segment.append(indices[i])
            else:
                segments.append(current_segment)
                current_segment = [indices[i]]
        
        segments.append(current_segment)  # Add last segment
        return segments

    # ------------------------------------------------------
    # Unified Sentence Grouping Logic (3-Pass Strategy)
    # ------------------------------------------------------
    def build_grouped_lines(self, call_df, call_type="Sales Call", include_system_audio=False):
        """
        Unified grouped lines builder with configurable call type support.
        
        - call_type="Sales Call": Detects and excludes system audio recordings for grouping
        - call_type="SQCCB": Processes all rows without system recording detection
        - Uses three-pass physical merge strategy for optimal context grouping
        - Pass A: Forward merge for non-punctuated sentences
        - Pass B: Backward merge for short utterances  
        - Pass C: Forward merge for contextual windows
        """
        speaker_col, text_col = 'Speaker Roles', 'Transcription'
        start_col, end_col = 'Segment Start Time', 'Segment End Time'
        
        if speaker_col not in call_df.columns or text_col not in call_df.columns:
            raise KeyError(f"Required columns '{speaker_col}' and '{text_col}' not found.")
        
        # Create working copy to avoid modifying original DataFrame
        updated_call_df = call_df.copy()
        
        # Conditional system recording detection based on call_type
        if call_type == "Sales Call":
            # Sales Call mode: Detect system recordings
            system_recording_flags = self.detect_system_audio_recordings(call_df)
            updated_call_df['System_Audio_Recording'] = system_recording_flags.map({True: 'YES', False: 'NO'})
            
            # Apply system audio inclusion switch
            if include_system_audio:
                print(f"📊 Sales Call: Including all {len(call_df)} rows (including {system_recording_flags.sum()} system recordings) in analysis")
                analysis_df = updated_call_df.copy()
            else:
                print(f"📊 Sales Call: Filtering out {system_recording_flags.sum()} system recording rows from analysis")
                analysis_df = updated_call_df[~system_recording_flags].copy()
                print(f"📊 Sales Call: {len(analysis_df)} rows remaining for coverage analysis")
        else:
            # SQCCB mode: Skip system recording detection, process all rows
            print(f"🔍 SQCCB mode: Skip system recording detection, processing all {len(call_df)} rows")
            system_recording_flags = pd.Series([False] * len(call_df), index=call_df.index)
            updated_call_df['System_Audio_Recording'] = 'NO'
            analysis_df = updated_call_df.copy()
        
        def ends_with_punct(s):
            return s and s[-1] in '。！？；，.!?;'
        
        initial_rows = []
        for idx, row in analysis_df.iterrows():  # Use filtered DataFrame
            text = str(row.get(text_col, '')).strip()
            if not text: continue
            initial_rows.append({
                'orig_indices': [idx], 'text': text, 'speaker': row.get(speaker_col),
                'start_time': row.get(start_col), 'end_time': row.get(end_col),
            })
        
        if not initial_rows:
            print(f"⚠️  {call_type}: No valid rows found for grouping")
            return [], {}, updated_call_df, system_recording_flags
        
        # --- Pass A: Forward merge (non-punctuated) ---
        pass_a_results = []
        if initial_rows:
            pass_a_results.append(initial_rows[0].copy())
            for i in range(1, len(initial_rows)):
                current_row, last_group = initial_rows[i], pass_a_results[-1]
                if (last_group['speaker'] == current_row['speaker'] and
                    not ends_with_punct(last_group['text']) and
                    (len(last_group['text']) + len(current_row['text']) + 1) <= 150):
                    last_group['text'] += ' ' + current_row['text']
                    last_group['orig_indices'].extend(current_row['orig_indices'])
                    if pd.notna(current_row['end_time']): last_group['end_time'] = current_row['end_time']
                else:
                    pass_a_results.append(current_row.copy())
        
        # --- Pass B: Backward merge (short utterances) ---
        pass_b_results = []
        if pass_a_results:
            for row in reversed(pass_a_results):
                if (pass_b_results and 
                    pass_b_results[0]['speaker'] == row['speaker'] and
                    len(pass_b_results[0]['text']) < 20 and
                    (len(row['text']) + len(pass_b_results[0]['text']) + 1) <= 150):
                    first_group = pass_b_results[0]
                    first_group['text'] = row['text'] + ' ' + first_group['text']
                    first_group['orig_indices'] = row['orig_indices'] + first_group['orig_indices']
                    if pd.notna(row['start_time']): first_group['start_time'] = row['start_time']
                else:
                    pass_b_results.insert(0, row.copy())
        
        # --- Pass C: Forward contextual window merge ---
        final_results = []
        if pass_b_results:
            context_window_limit = 150  # Keep consistent with other limits
            final_results.append(pass_b_results[0].copy())
            for i in range(1, len(pass_b_results)):
                current_row, last_group = pass_b_results[i], final_results[-1]
                if (last_group['speaker'] == current_row['speaker'] and
                    (len(last_group['text']) + len(current_row['text']) + 1) <= context_window_limit):
                    last_group['text'] += ' ' + current_row['text']
                    last_group['orig_indices'].extend(current_row['orig_indices'])
                    if pd.notna(current_row['end_time']): last_group['end_time'] = current_row['end_time']
                else:
                    final_results.append(current_row.copy())
        
        # --- Final formatting ---
        grouped_lines, original_to_group = [], {}
        for group_id_counter, group in enumerate(final_results, 1):
            start_idx, end_idx = min(group['orig_indices']), max(group['orig_indices'])
            grouped_lines.append({
                'group_id': group_id_counter, 'text': group['text'], 'speaker': group['speaker'],
                'start_idx': start_idx, 'end_idx': end_idx,
                'start_time': group['start_time'], 'end_time': group['end_time'],
                'orig_indices': group['orig_indices']  # ✅ 添加orig_indices字段
            })
            for orig_idx in group['orig_indices']:
                original_to_group[orig_idx] = group_id_counter
        
        print(f"🔀 {call_type} grouping completed: {len(grouped_lines)} final groups created")
        
        # Return enhanced results: grouped_lines, original_to_group, updated_call_df, system_recording_flags
        return grouped_lines, original_to_group, updated_call_df, system_recording_flags

    # ------------------------------------------------------
    # Legacy Sentence Grouping Logic (3-Pass Strategy)
    # ------------------------------------------------------

    
    # ------------------------------------------------------
    # Helper & Pattern Recognition Functions
    # ------------------------------------------------------
    def get_business_overlapping_keywords(self, text1, text2):
        """
        Find overlapping business keywords between two texts.
        Prioritizes important keywords first.
        """
        tokens1 = set(self.tokenize_text(text1))
        tokens2 = set(self.tokenize_text(text2))
        all_overlap = tokens1 & tokens2
        
        # Separate important and regular overlapping keywords
        important_overlap = all_overlap & self.important_keywords
        regular_overlap = all_overlap - important_overlap
        
        # Return important keywords first, then regular ones
        return list(important_overlap) + list(regular_overlap)

    def calculate_keyword_coverage(self, script_text, group_text):
        """
        Calculate keyword coverage ratio for business-critical terms.
        """
        # 改动点1: 使用缓存版本避免重复预处理
        script_tokens = set(self.tokenize_text(self._preprocess_memo(script_text)))
        group_tokens = set(self.tokenize_text(self._preprocess_memo(group_text)))
        
        # Focus on important keywords from script
        important_script_keywords = script_tokens & self.important_keywords
        if not important_script_keywords:
            return 0.0
        
        # Calculate coverage of important keywords
        covered_keywords = group_tokens & important_script_keywords
        return float(len(covered_keywords)) / len(important_script_keywords)

    def setup_enhanced_scoring_config(self):
        """
        Load enhanced scoring configuration from USER_ENHANCED_DISCUSSION_POINTS defined at top of file.
        
        Returns the configuration dictionary for pattern-based score enhancements.
        
        Usage: Configure discussion points at the top of this file by modifying the
        USER_ENHANCED_DISCUSSION_POINTS dictionary.
        """
        return USER_ENHANCED_DISCUSSION_POINTS
    
    def detect_date_patterns(self, text):
        """
        Detect various Chinese date patterns in text including mixed Arabic/Chinese numerals.
        
        Supports formats like:
        - 2025年10月9日
        - 二零二五年十月九日  
        - 10月9日
        - 十月九日
        - 2025年十月九日 (mixed)
        - 二零二五年10月9日 (mixed)
        - 2025-10-09, 2025/10/09
        """
        # Quick check for date indicators
        if not any(indicator in text for indicator in ['年', '月', '日', '/', '-']):
            return False, []
        
        # Comprehensive date patterns for Chinese dates
        patterns = [
            # Full Chinese dates with years, months, days
            r'[二零一三四五六七八九]{2,4}年[一二三四五六七八九十\d]{1,3}月[一二三四五六七八九十\d]{1,3}日',
            # Mixed Arabic year with Chinese month/day
            r'\d{4}年[一二三四五六七八九十\d]{1,3}月[一二三四五六七八九十\d]{1,3}日',
            # Chinese year with Arabic month/day  
            r'[二零一三四五六七八九]{2,4}年\d{1,2}月\d{1,2}日',
            # Standard Arabic format
            r'\d{4}年\d{1,2}月\d{1,2}日',
            # Month-day only formats
            r'[一二三四五六七八九十\d]{1,3}月[一二三四五六七八九十\d]{1,3}日',
            r'\d{1,2}月\d{1,2}日',
            # Western date formats
            r'\d{4}[-/]\d{1,2}[-/]\d{1,2}'
        ]
        
        found_dates = []
        for pattern in patterns:
            matches = re.findall(pattern, text)
            found_dates.extend(matches)
        
        return len(found_dates) > 0, found_dates[:3]  # Return first 3 matches
    
    def detect_numeric_patterns(self, text):
        """
        Detect financial/numeric patterns while avoiding date-like sequences.
        
        Looks for:
        - Monetary amounts (元, 万, 千, 亿)
        - Currency symbols (¥, $, €)
        - Percentages
        - Large numbers
        """
        # Quick check for numeric indicators
        if not re.search(r'[\d¥$€%万千百亿元]', text):
            return False, []
        
        patterns = [
            # Chinese monetary amounts  
            r'[港人民币港币]{1,3}\s*\d+(?:\.\d+)?[万千百亿元]?',
            # Currency symbols with amounts
            r'[¥$€]\s*[\d,]+(?:\.\d+)?',
            # Standalone amounts with Chinese units
            r'\d+(?:\.\d+)?[万千百亿元]',
            # Percentages
            r'\d+(?:\.\d+)?%'
        ]
        
        found_numbers = []
        for pattern in patterns:
            matches = re.findall(pattern, text)
            found_numbers.extend(matches)
        
        return len(found_numbers) > 0, found_numbers[:2]  # Return first 2 matches
    
    def apply_pattern_enhancement(self, base_score, text, discussion_point, dynamic_numeric_boost=0):
        """
        Apply pattern-based score enhancement for specific discussion points.
        This is called automatically during similarity calculation.
        
        How it works:
        1. Checks if the discussion_point is configured for enhancement
        2. If configured, detects date/numeric patterns in the text
        3. Applies the configured boost weights if patterns are found
        4. If no static config, applies dynamic numeric boost if available
        5. Returns the enhanced score (capped at 1.0)
        
        Configuration: Modify USER_ENHANCED_DISCUSSION_POINTS at the top of this file
        to specify which discussion points should receive which type of boosts.
        
        Weight Guidelines:
        - date_boost: 0.05-0.25 for discussion points about dates/deadlines
        - numeric_boost: 0.05-0.25 for discussion points about prices/amounts
        - dynamic_numeric_boost: 0.15 for auto-detected price/interest related points (English keywords only)
        
        Efficiency: The system efficiently skips pattern detection for discussion
        points that are not configured for enhancement, saving computation time.
        """
        enhancement_config = self.setup_enhanced_scoring_config()
        enhanced_score = base_score
        
        # Priority 1: Use static configuration if available
        if discussion_point in enhancement_config:
            config = enhancement_config[discussion_point]

            # Apply date pattern boost if configured
            if config.get('date_boost', 0) > 0:
                has_dates, _ = self.detect_date_patterns(text)
                if has_dates:
                    enhanced_score += config['date_boost']

            # Apply numeric pattern boost if configured  
            if config.get('numeric_boost', 0) > 0:
                has_numbers, _ = self.detect_numeric_patterns(text)
                if has_numbers:
                    enhanced_score += config['numeric_boost']
        
        # Priority 2: Use dynamic numeric boost if no static config and dynamic boost is available
        elif dynamic_numeric_boost > 0:
            has_numbers, _ = self.detect_numeric_patterns(text)
            if has_numbers:
                enhanced_score += dynamic_numeric_boost

        return min(1.0, enhanced_score)  # Cap at 1.0

    def calculate_token_rouge_l(self, text1, text2):
        """
        Calculate ROUGE-L (Longest Common Subsequence) F-score at character level.
        """
        seq1 = list(text1)
        seq2 = list(text2)
        
        if not seq1 or not seq2:
            return 0.0
        
        n, m = len(seq1), len(seq2)
        dp = [[0] * (m + 1) for _ in range(n + 1)]
        
        # Dynamic programming to find LCS length
        for i in range(1, n + 1):
            for j in range(1, m + 1):
                if seq1[i-1] == seq2[j-1]:
                    dp[i][j] = dp[i-1][j-1] + 1
                else:
                    dp[i][j] = max(dp[i-1][j], dp[i][j-1])
        
        lcs_len = dp[n][m]
        if lcs_len == 0:
            return 0.0
        
        # Calculate precision, recall, and F-score
        precision = lcs_len / m
        recall = lcs_len / n
        
        if precision + recall == 0:
            return 0.0
        
        return (2 * precision * recall) / (precision + recall)

    # ------------------------------------------------------
    # Core Semantic Similarity Calculation (UPGRADED TO SBERT)
    # ------------------------------------------------------
    def calculate_semantic_similarity(self, text1, text2, group_text_to_embedding=None):
        """
        Calculates semantic similarity using a hybrid system led by SBERT.
        
        Metrics Breakdown:
        1. SBERT Semantic Score (60%): Deep semantic understanding using transformer embeddings
        2. Expanded Token Overlap (15%): Synonym-aware matching for business terms
        3. ROUGE-L (15%): Sequence similarity (LCS-based character matching)
        4. Keyword Coverage (10%): Compliance check for critical business terms
        
        Returns:
            dict: Dictionary containing all individual metrics and the final weighted score
        """
        # 改动点1: 使用缓存版本避免重复预处理
        clean1 = self._preprocess_memo(text1)
        clean2 = self._preprocess_memo(text2)

        # Use cache to avoid re-computing for semantically identical pairs
        cache_key = (clean1, clean2)
        if cache_key in self._similarity_cache:
            return self._similarity_cache[cache_key]
        metrics = {}
        
        # 1. SBERT Semantic Score (Optimized with precomputed embeddings)
        try:
            # Try to get precomputed embeddings first
            embedding1 = None
            embedding2 = None
            
            # Check for script embedding (normalized clean-text key)
            if self.script_embeddings and clean1 in self.script_embeddings:
                embedding1 = self.script_embeddings[clean1]
            
            # Check for call text embedding (normalized clean-text key)
            if group_text_to_embedding and clean2 in group_text_to_embedding:
                embedding2 = group_text_to_embedding[clean2]
            
            # Fall back to on-demand encoding if needed
            if embedding1 is None and embedding2 is None:
                # Both need encoding
                embeddings = self.sbert_model.encode([clean1, clean2])
                embedding1, embedding2 = embeddings[0], embeddings[1]
            elif embedding1 is None:
                # Only script text needs encoding
                embedding1 = self.sbert_model.encode([clean1])[0]
            elif embedding2 is None:
                # Only call text needs encoding
                embedding2 = self.sbert_model.encode([clean2])[0]
            
            # Calculate cosine similarity
            semantic_score = util.cos_sim(embedding1, embedding2)[0][0].item()
            
        except Exception as e:
            print(f"⚠️  SBERT similarity calculation failed for texts: '{clean1[:20]}...' | '{clean2[:20]}...'. Error: {e}. Falling back to 0.0")
            semantic_score = 0.0
        metrics['semantic_score'] = semantic_score
        
        # 2. Expanded token overlap (with synonyms) - Still valuable for business logic
        # 改动点1: 使用缓存版本避免重复关键词扩展
        expanded1 = self._expand_keywords_memo(clean1)
        expanded2 = self._expand_keywords_memo(clean2)
        union_size = len(expanded1 | expanded2)
        expanded_similarity = len(expanded1 & expanded2) / union_size if union_size > 0 else 0
        metrics['expanded_overlap'] = expanded_similarity
        
        # 3. Token-level ROUGE-L (LCS-based sequence similarity)
        rouge_l = self.calculate_token_rouge_l(clean1, clean2)
        metrics['rouge_l'] = rouge_l

        # 4. Keyword coverage (bounded 0..1)
        keyword_coverage = self.calculate_keyword_coverage(text1, text2)
        metrics['keyword_coverage'] = keyword_coverage
        
        # New weighted combination with language-specific weights from config
        try:
            # 统一从config读取权重，移除硬编码
            language = getattr(self, 'current_language', 'MAN')  # 默认MAN确保向后兼容
            from config import get_similarity_weights
            weights = get_similarity_weights(language)
            
            weighted_score = (
                semantic_score * weights.get('SBERT', 0.60) +
                expanded_similarity * weights.get('expanded_overlap', 0.15) +
                rouge_l * weights.get('rouge_l', 0.15) +
                keyword_coverage * weights.get('keyword_coverage', 0.10)
            )
            
            # ENG路径日志
            if language == 'ENG':
                print(f"🔤 ENG similarity calculation: SBERT={semantic_score:.3f}, expanded={expanded_similarity:.3f}, rouge_l={rouge_l:.3f}, keyword={keyword_coverage:.3f}, weighted={weighted_score:.3f}")
                
        except Exception as e:
            # 异常回退到MAN默认权重，确保稳定性
            print(f"⚠️  Error loading similarity weights: {e}, using MAN defaults")
            weighted_score = (
                semantic_score * 0.60 +          # SBERT provides the main semantic signal
                expanded_similarity * 0.15 +     # Synonyms provide explicit business logic
                rouge_l * 0.15 +                 # Sequence matching guards against nonsensical order
                keyword_coverage * 0.10          # Critical keywords act as a compliance backstop
            )
        metrics['weighted_score'] = weighted_score
        
        # Update cache (limit size to prevent memory bloat)
        if len(self._similarity_cache) < 5000:
            self._similarity_cache[cache_key] = metrics

        return metrics
    
    # ------------------------------------------------------
    # Script & Coverage Logic (Adapted for SBERT)
    # ------------------------------------------------------
    def parse_script_variations(self, script_text):
        """
        Parse script variations from text, handling multiple formats and versions.
        Enhanced with preprocessing and comprehensive short fragment merging.
        
        Args:
            script_text (str): Raw script text that may contain multiple variations
            
        Returns:
            list: List of individual script variations
        """
        if pd.isna(script_text):
            return []
        
        script_text = str(script_text).strip()
        if not script_text:
            return []
        
        # Step 1: 先预处理去英文
        preprocessed_text = self.preprocess_text(script_text, mode='comparison')
        if not preprocessed_text:
            return []
        
        # Step 2: 第一次分割
        raw_splits = re.split(r'[.;\n]|版本[AB][:：]|Version [AB][:：]|情况[一二三四五六七八九十][:：]', preprocessed_text)
        
        # Step 3: 第一次短片段合并
        cleaned_splits = [split.strip() for split in raw_splits if split.strip()]
        merged_splits = self._merge_short_fragments(cleaned_splits)
        
        # Step 4: 第二次分割（处理过长片段）
        variations = []
        for split in merged_splits:
            if len(split) > 50:
                # 按逗号分割
                comma_splits = [s.strip() for s in split.split(',') if s.strip()]
                # 第二次短片段合并
                comma_merged = self._merge_short_fragments(comma_splits)
                variations.extend(comma_merged)
            else:
                variations.append(split)
        
        # Step 5: 去重
        unique_variations = []
        seen = set()
        for var in variations:
            if var not in seen:
                unique_variations.append(var)
                seen.add(var)
        
        return unique_variations if unique_variations else [preprocessed_text]
    
    def _merge_short_fragments(self, fragments):
        """
        Helper function: 合并≤5字的短片段
        短片段向后合并，最后一个向前合并
        
        Args:
            fragments: List of text fragments
            
        Returns:
            List of merged fragments with no fragment ≤5 characters
        """
        if not fragments:
            return []
        
        merged = []
        i = 0
        while i < len(fragments):
            current = fragments[i]
            
            if len(current) <= 5:
                # 尝试向后合并
                if i + 1 < len(fragments):
                    fragments[i + 1] = current + ' ' + fragments[i + 1]
                # 如果是最后一个，向前合并
                elif merged:
                    merged[-1] = merged[-1] + ' ' + current
                else:
                    merged.append(current)  # 只有这一个片段
            else:
                merged.append(current)
            i += 1
        
        return merged

    def compute_pairwise_matches(self, grouped_lines, script_df, group_text_to_embedding=None, *, threshold=0.4, 
                                weights=None, settings=None, enable_topk_filter=False, topk=3):
        """
        核心计算函数：统一计算所有(对话组, 讨论点)配对的全部指标，实现"一次计算，多处复用"。
        
        职责：
        - 对全部(group, point)配对计算SBERT、ROUGE-L、关键词覆盖、动态数值加成、加权分数
        - 复用现有批量编码结果和缓存机制，避免重复计算
        - 产出中间结果表，供Coverage Analysis和Call Text Analysis共同使用
        
        Args:
            grouped_lines: 对话组列表
            script_df: 脚本DataFrame
            group_text_to_embedding: 预计算的向量映射
            threshold: SBERT门控阈值（默认0.4）
            weights: 权重配置（保留扩展性）
            settings: 设置配置（保留扩展性）
            enable_topk_filter: 是否启用top-k过滤（默认False）
            topk: top-k数量（默认3）
            
        Returns:
            pd.DataFrame: 配对结果表，列包含group_id, group_text, point_id, point_text, 
                         sbert_similarity, rouge_l, keyword_coverage, expanded_overlap,
                         dynamic_numeric_boost, weighted_score, match_type, best_variation,
                         holistic_score, granular_score等
        """
        pairwise_results = []
        
        # Group scripts by discussion point and create holistic scripts
        point_to_scripts = {}
        point_to_variations = {}
        point_to_complete_holistic = {}
        
        for _, row in script_df.iterrows():
            point = row['Required_Discussion_Point']
            script = row['Standard_Script']
            if pd.notna(point) and pd.notna(script):
                if point not in point_to_scripts:
                    point_to_scripts[point] = []
                    point_to_variations[point] = []
                    point_to_complete_holistic[point] = []
                
                point_to_scripts[point].append(str(script))
                point_to_variations[point].extend(self.parse_script_variations(str(script)))
                point_to_complete_holistic[point].append(str(script))
        
        # Create complete holistic script text for each point
        for point in point_to_complete_holistic:
            point_to_complete_holistic[point] = ' '.join(point_to_complete_holistic[point])
        
        # 遍历所有(group, point)配对
        for group_info in grouped_lines:
            group_text = group_info['text']
            group_id = group_info.get('group_id', '')
            
            if not group_text:
                continue
                
            for point, scripts in point_to_scripts.items():
                # Dynamic numeric boost detection
                dynamic_numeric_boost = 0
                point_lower = point.lower()
                if 'price' in point_lower or 'floating rate' in point_lower:
                    dynamic_numeric_boost = 0.15
                elif dynamic_numeric_boost == 0:
                    holistic_script = point_to_complete_holistic.get(point, "")
                    if holistic_script:
                        combined_scripts_lower = holistic_script.lower()
                        if any(keyword in combined_scripts_lower for keyword in ['price', 'interest', '%']):
                            dynamic_numeric_boost = 0.15
                
                # Step 1: Holistic matching - 复用现有逻辑
                holistic_script_text = point_to_complete_holistic.get(point, "")
                script_variations = point_to_variations.get(point, [])
                
                holistic_metrics = self.calculate_semantic_similarity(
                    holistic_script_text, group_text, group_text_to_embedding
                )
                holistic_weighted_score = holistic_metrics['weighted_score']
                
                # 改动点3: Holistic优先早停 - 先计算增强分数，达到阈值则跳过Granular扫描
                enhanced_holistic_score = self.apply_pattern_enhancement(
                    holistic_weighted_score, group_text, point, dynamic_numeric_boost
                )
                
                # 如果Holistic已达到阈值，直接选择，避免Granular循环
                if enhanced_holistic_score >= threshold:
                    # 早停：直接选择Holistic作为胜出者
                    final_score = enhanced_holistic_score
                    final_metrics = holistic_metrics
                    final_variation = holistic_script_text
                    final_match_type = "Holistic"
                    final_holistic_score = enhanced_holistic_score
                    final_granular_score = 0  # 未执行granular
                else:
                    # Step 2: Granular matching - 只有当Holistic未达阈值时才执行
                    best_granular_score_for_group = 0
                    best_granular_metrics = {}
                    best_granular_variation = ""
                    
                    for variation in script_variations:
                        metrics = self.calculate_semantic_similarity(
                            variation, group_text, group_text_to_embedding
                        )
                        granular_score = metrics['weighted_score']
                        
                        if granular_score > best_granular_score_for_group:
                            best_granular_score_for_group = granular_score
                            best_granular_metrics = metrics
                            best_granular_variation = variation
                    
                    # Apply pattern enhancement to best granular score
                    enhanced_granular_score = self.apply_pattern_enhancement(
                        best_granular_score_for_group, group_text, point, dynamic_numeric_boost
                    )
                    
                    # Step 3: Choose the better approach between Holistic and Granular
                    if enhanced_holistic_score > enhanced_granular_score:
                        final_score = enhanced_holistic_score
                        final_metrics = holistic_metrics
                        final_variation = holistic_script_text
                        final_match_type = "Holistic"
                        final_holistic_score = enhanced_holistic_score
                        final_granular_score = enhanced_granular_score
                    else:
                        final_score = enhanced_granular_score
                        final_metrics = best_granular_metrics
                        final_variation = best_granular_variation
                        final_match_type = "Granular"
                        final_holistic_score = enhanced_holistic_score
                        final_granular_score = enhanced_granular_score
                
                # 收集配对结果
                pairwise_result = {
                    'group_id': group_id,
                    'group_text': group_text,
                    'point_id': point,
                    'point_text': point,
                    'sbert_similarity': final_metrics.get('semantic_score', 0),
                    'rouge_l': final_metrics.get('rouge_l', 0),
                    'keyword_coverage': final_metrics.get('keyword_coverage', 0),
                    'expanded_overlap': final_metrics.get('expanded_overlap', 0),
                    'dynamic_numeric_boost': dynamic_numeric_boost,
                    'weighted_score': final_score,
                    'match_type': final_match_type,
                    'best_variation': final_variation,
                    'holistic_score': final_holistic_score,
                    'granular_score': final_granular_score,
                    'all_variations_count': len(point_to_variations.get(point, [])),
                    # 添加报告阶段需要的元数据
                    'speaker': group_info.get('speaker', ''),
                    'timestamp': group_info.get('timestamp', ''),
                    'start_time': group_info.get('start_time', ''),
                    'end_time': group_info.get('end_time', ''),
                }
                
                pairwise_results.append(pairwise_result)
        
        return pd.DataFrame(pairwise_results)

    def check_coverage(self, call_df, required_points_df, grouped_lines, threshold=0.4, group_text_to_embedding=None):
        """
        Coverage Analysis报告函数：复用核心计算函数，针对每个讨论点选择全通话最佳匹配。
        
        Args:
            call_df: DataFrame with call transcription data
            required_points_df: DataFrame with required discussion points and scripts
            grouped_lines: List of grouped call text segments
            threshold: Minimum similarity score to consider a point "covered"
            group_text_to_embedding: Dict mapping preprocessed call text to precomputed embeddings (optional)
            
        Returns:
            pd.DataFrame: Coverage analysis results（保持原有返回结构与列顺序）
        """
        # 调用核心计算函数，获取所有配对结果
        pairwise_df = self.compute_pairwise_matches(
            grouped_lines, required_points_df, group_text_to_embedding, 
            threshold=threshold
        )
        
        results = []
        
        # 针对每个讨论点，从配对结果中选择weighted_score最高的一行作为最佳匹配
        for point_id in pairwise_df['point_id'].unique():
            point_matches = pairwise_df[pairwise_df['point_id'] == point_id]
            
            if len(point_matches) == 0:
                continue
                
            # 选择weighted_score最高的匹配（保持现有tie-break规则）
            best_match = point_matches.loc[point_matches['weighted_score'].idxmax()]
            
            # 改动点2: 仅对最终胜出者做关键词重叠计算（复用memo缓存）
            if best_match['weighted_score'] > 0:
                overlapping_keywords = self.get_business_overlapping_keywords(
                    self._preprocess_memo(best_match['best_variation']),
                    best_match['group_text']
                )
            else:
                overlapping_keywords = []
            
            # 构建与原版本完全一致的返回结构
            result = {
                'Required_Discussion_Point': best_match['point_id'],
                'Covered': 'Covered' if best_match['weighted_score'] >= threshold else 'Not Covered',
                'Weighted_Score': round(best_match['weighted_score'], 3),
                'SBERT_Similarity': round(best_match['sbert_similarity'], 3),
                'ROUGE_L': round(best_match['rouge_l'], 3),
                'Keyword_Coverage': round(best_match['keyword_coverage'], 3),
                'Expanded_Overlap': round(best_match['expanded_overlap'], 3),
                'Dynamic_Numeric_Boost': round(best_match['dynamic_numeric_boost'], 3),
                'Match_Type': best_match['match_type'],
                'Best_Matching_Variation': self.preprocess_text(best_match['best_variation'], mode='comparison')[:100] + '...' if len(self.preprocess_text(best_match['best_variation'], mode='comparison')) > 100 else self.preprocess_text(best_match['best_variation'], mode='comparison'),
                'Matched_Group': best_match['group_text'][:200] + '...' if len(best_match['group_text']) > 200 else best_match['group_text'],
                'Group_ID': best_match['group_id'],
                'Speaker': best_match['speaker'],
                'Overlapping_Keywords': ', '.join(overlapping_keywords),
                'All_Variations_Count': best_match['all_variations_count'],
                'Holistic_Score': round(best_match['holistic_score'], 3),
                'Granular_Score': round(best_match['granular_score'], 3),
            }
            
            results.append(result)
        
        return pd.DataFrame(results)
    
    def _analyze_single_point_coverage(self, point, scripts, grouped_lines, threshold, group_text_to_embedding=None):
        """
        SBERT优化版本：采用双重验证策略
        
        IMPLEMENTS DUAL VERIFICATION: Holistic vs Granular matching strategy.
        
        策略：
        1. 解析脚本变体得到"精准小靶子"（Granular matching）
        2. 创建完整脚本文本用于整段匹配（Holistic matching）
        3. 对每个对话组进行双重匹配，取最高分
        4. 追踪匹配类型和分数用于诊断
        
        Args:
            point: Discussion point name
            scripts: List of standard scripts for this point
            grouped_lines: List of grouped call text segments (已包含丰富上下文)
            threshold: Coverage threshold
            
        Returns:
            dict: Analysis results for this discussion point
        """
        # 步骤1: 解析脚本变体，得到"精准小靶子"列表（用于Granular matching）
        script_variations = []
        for script in scripts:
            script_variations.extend(self.parse_script_variations(script))
        
        # 步骤2: 创建完整脚本文本（用于Holistic matching）
        holistic_script_text = ' '.join(scripts) if scripts else ""
        
        # Dynamic numeric boost detection
        dynamic_numeric_boost = 0
        
        # Dynamic rule 1: Check point name for price/floating rate keywords
        point_lower = point.lower()
        if 'price' in point_lower or 'floating rate' in point_lower:
            dynamic_numeric_boost = 0.15
        
        # Dynamic rule 2: Check script content for price/interest/% keywords
        if dynamic_numeric_boost == 0 and scripts:
            combined_scripts_lower = ' '.join(scripts).lower()
            if any(keyword in combined_scripts_lower for keyword in ['price', 'interest', '%']):
                dynamic_numeric_boost = 0.15
        
        # 初始化最佳匹配结果
        best_score = 0.0
        best_original_score = 0.0
        best_group_info = {}
        best_variation = ""
        best_metrics = {}
        best_match_type = ""  # Track whether final match was Holistic or Granular
        best_holistic_score = 0  # Track holistic score for the best match
        best_granular_score = 0  # Track granular score for the best match
        
        # Get point-specific weights for this discussion point
        try:
            point_weights = dictionaries.get_point_specific_weights(
                self.current_language, self.current_product, point
            )
            # Temporarily switch to point-specific weights
            original_weights = self.current_weights
            self.current_weights = point_weights
            point_specific_mode = True
        except Exception as e:
            # Fallback to existing weights if point-specific weights fail
            point_specific_mode = False
        
        # 步骤3: 遍历包含丰富上下文的对话组，进行双重验证
        for group_info in grouped_lines:
            group_text = group_info['text']
            if not group_text:
                continue
            
            # Step 1: Holistic matching - compare with complete script text
            holistic_metrics = self.calculate_semantic_similarity(holistic_script_text, group_text, group_text_to_embedding)
            holistic_weighted_score = holistic_metrics['weighted_score']
            
            # 改动点3: Holistic优先早停 - 先计算增强分数，达到阈值则跳过Granular扫描
            enhanced_holistic_score = self.apply_pattern_enhancement(
                holistic_weighted_score, group_text, point, dynamic_numeric_boost
            )
            
            # 如果Holistic已达到阈值，直接选择，避免Granular循环
            if enhanced_holistic_score >= threshold:
                # 早停：直接选择Holistic作为胜出者
                final_score = enhanced_holistic_score
                final_original_score = holistic_weighted_score
                final_metrics = holistic_metrics
                final_variation = holistic_script_text
                final_match_type = "Holistic"
                final_holistic_score = enhanced_holistic_score
                final_granular_score = 0  # 未执行granular
            else:
                # Step 2: Granular matching - find best among script variations
                best_granular_score_for_group = 0
                best_granular_metrics = {}
                best_granular_variation = ""
                
                for variation in script_variations:
                    metrics = self.calculate_semantic_similarity(variation, group_text, group_text_to_embedding)
                    granular_score = metrics['weighted_score']
                
                if granular_score > best_granular_score_for_group:
                    best_granular_score_for_group = granular_score
                    best_granular_metrics = metrics
                    best_granular_variation = variation
            
                # Apply pattern enhancement to best granular score
                enhanced_granular_score = self.apply_pattern_enhancement(
                    best_granular_score_for_group, group_text, point, dynamic_numeric_boost
                )
                
                # Step 3: Choose the better approach between Holistic and Granular
                if enhanced_holistic_score > enhanced_granular_score:
                    final_score = enhanced_holistic_score
                    final_original_score = holistic_weighted_score
                    final_metrics = holistic_metrics
                    final_variation = holistic_script_text
                    final_match_type = "Holistic"
                    final_holistic_score = enhanced_holistic_score
                    final_granular_score = enhanced_granular_score
                else:
                    final_score = enhanced_granular_score
                    final_original_score = best_granular_score_for_group
                    final_metrics = best_granular_metrics
                    final_variation = best_granular_variation
                    final_match_type = "Granular"
                    final_holistic_score = enhanced_holistic_score
                    final_granular_score = enhanced_granular_score
            
            # Update best match if this group scored higher
            if final_score > best_score:
                best_score = final_score
                best_original_score = final_original_score
                best_group_info = group_info
                best_metrics = final_metrics
                best_variation = final_variation
                best_match_type = final_match_type
                best_holistic_score = final_holistic_score
                best_granular_score = final_granular_score
                
                # Early exit for very high scores (SBERT can achieve higher scores)
                if best_score >= 0.95:
                    break
            
            if best_score >= 0.95:
                break
        
        # Restore original weights if we switched to point-specific mode
        if point_specific_mode:
            self.current_weights = original_weights
        
        # 改动点2: 关键词重叠仅在胜出者一次性执行
        # 只对最终胜出者计算关键词，避免在循环中重复计算
        if best_group_info:
            overlapping_keywords = self.get_business_overlapping_keywords(
                self._preprocess_memo(best_variation or (scripts[0] if scripts else "")),
                best_group_info.get('text', '')
            )
        else:
            overlapping_keywords = []
        
        enhancement_boost = best_score - best_original_score
        
        return {
            'Required_Discussion_Point': point,
            'Covered': 'Covered' if best_score >= threshold else 'Not Covered',
            'Weighted_Score': round(best_score, 3),
            'Original_Score': round(best_original_score, 3),
            'Enhancement_Boost': round(enhancement_boost, 3),
            'SBERT_Semantic_Score': round(best_metrics.get('semantic_score', 0), 3),  # New metric
            'Expanded_Overlap': round(best_metrics.get('expanded_overlap', 0), 3),
            'ROUGE_L': round(best_metrics.get('rouge_l', 0), 3),
            'Keyword_Coverage': round(best_metrics.get('keyword_coverage', 0), 3),
            'Overlapping_Keywords': ', '.join(overlapping_keywords),
            'Matched_Group': best_group_info.get('text', ''),
            'Group_ID': best_group_info.get('group_id', -1),
            'Speaker': best_group_info.get('speaker', ''),
            'Best_Matching_Variation': self.preprocess_text(best_variation or (scripts[0] if scripts else ''), mode='comparison'),
            'All_Variations_Count': len(script_variations),
            'Match_Type': best_match_type,  # New: Holistic or Granular
            'Holistic_Score': round(best_holistic_score, 3),  # New: holistic matching score
            'Granular_Score': round(best_granular_score, 3),  # New: granular matching score
        }
        
        return result

    # ------------------------------------------------------
    # Reporting & Output Functions
    # ------------------------------------------------------
    def create_grouped_call_dataframe(self, grouped_lines):
        """Create DataFrame from grouped lines for export."""
        return pd.DataFrame(grouped_lines)
    
    def create_call_text_analysis_view(self, grouped_lines, script_df, threshold=0.4, group_text_to_embedding=None):
        """
        Call Text Analysis报告函数：复用核心计算函数，针对每个对话组生成对所有讨论点的匹配视图。
        
        Args:
            grouped_lines: List of grouped call text segments
            script_df: DataFrame with script data
            threshold: Threshold for determining hits
            group_text_to_embedding: Dict mapping preprocessed call text to precomputed embeddings (optional)
            
        Returns:
            pd.DataFrame: Analysis view with columns for each discussion point（保持原有返回结构与列顺序）
        """
        # 调用核心计算函数，获取所有配对结果
        pairwise_df = self.compute_pairwise_matches(
            grouped_lines, script_df, group_text_to_embedding, 
            threshold=threshold
        )
        
        analysis_rows = []
        all_points = script_df['Required_Discussion_Point'].dropna().unique()
        
        # 针对每个对话组，从配对结果中选择对每个讨论点的最佳匹配
        for group in grouped_lines:
            group_id = group.get('group_id', '')
            
            analysis_row = {
                'Group_ID': group_id,
                'Speaker': group.get('speaker', ''),
                'Call_Text': group.get('text', '')
            }
            
            # 为每个讨论点添加列
            for point in all_points:
                # 从配对结果中找到当前组对当前点的匹配
                group_point_matches = pairwise_df[
                    (pairwise_df['group_id'] == group_id) & 
                    (pairwise_df['point_id'] == point)
                ]
                
                if len(group_point_matches) > 0:
                    # 选择最佳匹配（通常只有一行）
                    best_match = group_point_matches.iloc[0]
                    
                    best_score_for_point = best_match['weighted_score']
                    point_hit = best_score_for_point >= threshold
                    
                    # 改动点2: 仅对胜出者生成关键词，使用缓存版本避免重复分词
                    if best_score_for_point > 0:
                        best_keywords_for_point = self._jieba_tokens_memo(group.get('text', ''))[:10]
                    else:
                        best_keywords_for_point = []
                else:
                    # 如果没有匹配结果，设置默认值
                    point_hit = False
                    best_score_for_point = 0.0
                    best_keywords_for_point = []
                
                # Create shortened point name for column
                point_short = f"{point[:30]}{'...' if len(point) > 30 else ''}"
                
                # Add columns for this discussion point
                analysis_row[f"{point_short}_Hit"] = 'YES' if point_hit else 'NO'
                analysis_row[f"{point_short}_Score"] = round(best_score_for_point, 3)
                analysis_row[f"{point_short}_Keywords"] = ', '.join(best_keywords_for_point)
            
            analysis_rows.append(analysis_row)
        
        return pd.DataFrame(analysis_rows)

    def create_call_text_analysis_view_with_separation(self, grouped_lines, system_audio_df, script_df, threshold=0.4, group_text_to_embedding=None):
        """
        Call Text Analysis报告函数（with_separation版本）：复用核心计算函数，并通过元数据进行人工对话与系统录音分离。
        
        This function implements the late merge strategy:
        1. Generate analysis results for human dialogue groups using shared pairwise computation
        2. Format system recording rows with minimal structure (_Hit='NO' only, no Score/Keywords)
        3. Combine both into a single comprehensive report（保持原有返回结构与列顺序）
        
        Args:
            grouped_lines: List of grouped call text segments  
            system_audio_df: DataFrame with system recording rows
            script_df: DataFrame with script data
            threshold: Threshold for determining hits
            group_text_to_embedding: Dict mapping preprocessed call text to precomputed embeddings (optional)
        """
        # 调用核心计算函数，获取所有配对结果（复用与其他报告函数相同的计算）
        pairwise_df = self.compute_pairwise_matches(
            grouped_lines, script_df, group_text_to_embedding, 
            threshold=threshold
        )
        
        analysis_rows = []
        all_points = script_df['Required_Discussion_Point'].dropna().unique()
        
        # Part 1: Process human dialogue groups using pairwise results
        for group in grouped_lines:
            group_id = group.get('group_id', '')
            
            analysis_row = {
                'Group_ID': group_id, 
                'Speaker': group.get('speaker', ''), 
                'Call_Text': group.get('text', ''),
                'System_Audio_Recording': 'NO'
            }
            
            # 为每个讨论点添加列（从pairwise_df获取结果）
            for point in all_points:
                point_short = f"{point[:30]}{'...' if len(point) > 30 else ''}"
                
                # 从配对结果中找到当前组对当前点的匹配
                group_point_matches = pairwise_df[
                    (pairwise_df['group_id'] == group_id) & 
                    (pairwise_df['point_id'] == point)
                ]
                
                if len(group_point_matches) > 0:
                    best_match = group_point_matches.iloc[0]
                    best_score_for_point = best_match['weighted_score']
                    point_hit = best_score_for_point >= threshold
                    
                    # 改动点2: 仅对胜出者生成Top 10关键词，使用缓存版本避免重复分词
                    if best_score_for_point > 0:
                        keywords = self._jieba_tokens_memo(group.get('text', ''))
                        word_freq = Counter(keywords)
                        best_keywords_for_point = [word for word, count in word_freq.most_common(10)]
                    else:
                        best_keywords_for_point = []
                else:
                    best_score_for_point = 0.0
                    point_hit = False
                    best_keywords_for_point = []
                
                # Record results for this point
                analysis_row[f"{point_short}_Hit"] = 'YES' if point_hit else 'NO'
                analysis_row[f"{point_short}_Score"] = round(best_score_for_point, 3)
                analysis_row[f"{point_short}_Keywords"] = ', '.join(best_keywords_for_point)
            
            analysis_rows.append(analysis_row)
        
        # Part 2: Process system recording rows (only Hit column needed)
        if not system_audio_df.empty:
            for idx, row in system_audio_df.iterrows():
                system_row = {
                    'Group_ID': f"SYS_{idx}",
                    'Speaker': row.get('Speaker Roles', 'System'),
                    'Call_Text': str(row.get('Transcription', '')),
                    'System_Audio_Recording': 'YES'
                }
                
                # For system recordings, only Hit column is needed
                for point in all_points:
                    point_short = f"{point[:30]}{'...' if len(point) > 30 else ''}"
                    system_row[f"{point_short}_Hit"] = 'NO'
                
                analysis_rows.append(system_row)
        
        return pd.DataFrame(analysis_rows)

    def create_sentence_level_output(self, call_df, original_to_group):
        """
        Create sentence-level output showing original to group mapping.
        
        Args:
            call_df: Original call DataFrame
            original_to_group: Mapping from original row indices to group IDs
            
        Returns:
            pd.DataFrame: Sentence-level analysis
        """
        speaker_col = 'Speaker Roles'
        text_col = 'Transcription'
        start_col = 'Segment Start Time'
        end_col = 'Segment End Time'
        
        sentence_level_data = []
        
        for idx, row in call_df.iterrows():
            text_val = row.get(text_col)
            if pd.isna(text_val) or not str(text_val).strip():
                continue
                
            sentence_level_data.append({
                'Original_Row': idx,
                'Group_ID': original_to_group.get(idx),
                'Speaker': row.get(speaker_col, ''),
                'Start_Time': row.get(start_col),
                'End_Time': row.get(end_col),
                'Original_Text': str(text_val).strip()
            })
        
        return pd.DataFrame(sentence_level_data)

# Old main function removed - using new four-step strategy below


# ==========================================================
# Executive Summary Report Generation
# ==========================================================

def parse_time_to_seconds(time_str):
    """
    Parse time string in format "00:18:00" to total seconds.
    Extracts the last 6 characters (MM:SS) and converts to seconds.
    """
    if pd.isna(time_str) or not time_str:
        return 0
    
    try:
        time_str = str(time_str).strip()
        # Extract last 6 characters for MM:SS format
        if len(time_str) >= 6:
            time_part = time_str[-6:]  # Get "18:00" from "00:18:00"
        else:
            time_part = time_str
        
        # Split and convert to seconds
        if ':' in time_part:
            parts = time_part.split(':')
            if len(parts) == 2:
                minutes = int(parts[0])
                seconds = int(parts[1])
                return minutes * 60 + seconds
            elif len(parts) == 3:  # Handle full HH:MM:SS
                hours = int(parts[0])
                minutes = int(parts[1])
                seconds = int(parts[2])
                return hours * 3600 + minutes * 60 + seconds
        
        # If no colon, try to parse as plain number
        return float(time_str)
    except (ValueError, IndexError):
        print(f"⚠️ Warning: Could not parse time format: {time_str}")
        return 0


def apply_composite_point_key_if_configured(script_df, script_sheet_name):
    """
    For configured sheets only, combine:
      Required_Discussion_Point + Points to Note
    as the new Required_Discussion_Point.
    """
    if not should_use_composite_point_key(script_sheet_name):
        return script_df

    required_col = 'Required_Discussion_Point'
    if required_col not in script_df.columns:
        print(f"⚠️  {script_sheet_name}: Missing '{required_col}', skip composite point key")
        return script_df

    note_col = None
    for col in script_df.columns:
        if str(col).strip().lower() == str(COMPOSITE_POINT_NOTE_COLUMN).strip().lower():
            note_col = col
            break

    if note_col is None:
        print(f"⚠️  {script_sheet_name}: Missing '{COMPOSITE_POINT_NOTE_COLUMN}', skip composite point key")
        return script_df

    updated_df = script_df.copy()

    def _combine(row):
        base = '' if pd.isna(row.get(required_col)) else str(row.get(required_col)).strip()
        note = '' if pd.isna(row.get(note_col)) else str(row.get(note_col)).strip()
        if note:
            return f"{base}{COMPOSITE_POINT_KEY_SEPARATOR}{note}" if base else note
        return base

    updated_df[required_col] = updated_df.apply(_combine, axis=1)
    print(f"🧩 {script_sheet_name}: Composite point key enabled ({required_col} + {note_col})")
    return updated_df

def generate_executive_summary_report(results_df, updated_call_df, script_df, grouped_lines, output_file=None):
    """
    Generate a consolidated Executive Summary Excel report with three sections:
    A: Overall Performance KPIs
    B: Key Insights - Speaker View  
    C: Detailed Coverage List
    
    Returns the final DataFrame that was written to Excel.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from collections import Counter
    import string
    
    print("📊 Generating Executive Summary report...")
    
    # ==========================================
    # Section A: Overall Performance KPIs
    # ==========================================
    print("📈 Calculating Overall Performance KPIs...")
    
    # Get total call duration directly from the last row's segment end time
    if 'Segment End Time' in updated_call_df.columns and not updated_call_df.empty:
        # Get the last row's end time value directly (keep as string if it's string)
        total_duration = updated_call_df['Segment End Time'].iloc[-1]
        if pd.isna(total_duration):
            total_duration = "N/A"
        else:
            total_duration = str(total_duration)
    else:
        total_duration = "N/A"
    
    # Calculate coverage metrics
    total_points = len(results_df)
    covered_points = len(results_df[results_df['Covered'] == 'Covered'])
    uncovered_points = total_points - covered_points
    coverage_rate = (covered_points / total_points * 100) if total_points > 0 else 0
    
    kpis_data = {
        'Metric': [
            'Total Call Duration',
            'Compliance Coverage Rate (%)',
            'Total Points Checked',
            'Covered Points',
            'Uncovered Risk Points'
        ],
        'Value': [
            str(total_duration),
            f"{coverage_rate:.1f}%",
            str(total_points),
            str(covered_points),
            str(uncovered_points)
        ]
    }
    kpis_df = pd.DataFrame(kpis_data)
    
    # ==========================================
    # Section B: Key Insights - Speaker View
    # ==========================================
    print("🎤 Generating Speaker View insights...")
    
    # Filter out system recordings for speaker analysis
    non_system_df = updated_call_df[updated_call_df.get('System_Audio_Recording', 'NO') == 'NO'].copy()
    
    # Calculate speaker statistics
    speaker_stats = []
    if not non_system_df.empty and 'Speaker Roles' in non_system_df.columns and 'Transcription' in non_system_df.columns:
        # Group by speaker and calculate metrics
        for speaker in non_system_df['Speaker Roles'].unique():
            if pd.isna(speaker):
                continue
                
            speaker_data = non_system_df[non_system_df['Speaker Roles'] == speaker]
            
            # Calculate speaking duration using proper time parsing with strptime
            if 'Segment Start Time' in speaker_data.columns and 'Segment End Time' in speaker_data.columns:
                speaking_duration = 0.0
                
                for _, segment in speaker_data.iterrows():
                    start_time_str = str(segment.get('Segment Start Time', '')).strip()
                    end_time_str = str(segment.get('Segment End Time', '')).strip()
                    
                    if start_time_str and end_time_str and start_time_str != 'nan' and end_time_str != 'nan':
                        try:
                            from datetime import datetime
                            # Parse time strings as HH:MM:SS format
                            start_time = datetime.strptime(start_time_str, "%H:%M:%S")
                            end_time = datetime.strptime(end_time_str, "%H:%M:%S")
                            
                            # Calculate duration in seconds
                            duration_seconds = (end_time - start_time).total_seconds()
                            if duration_seconds >= 0:  # Only add positive durations
                                speaking_duration += duration_seconds
                                
                        except ValueError:
                            # If parsing fails, try to use the old parse_time_to_seconds as fallback
                            try:
                                start_seconds = parse_time_to_seconds(start_time_str)
                                end_seconds = parse_time_to_seconds(end_time_str)
                                duration_seconds = end_seconds - start_seconds
                                if duration_seconds >= 0:
                                    speaking_duration += duration_seconds
                            except:
                                continue  # Skip this segment if can't parse
                
                print(f"🔍 Debug - Speaker: {speaker}, Speaking Duration: {speaking_duration:.1f}s")
            else:
                speaking_duration = 0.0
                print(f"🔍 Debug - Speaker: {speaker} - Time columns not found")
            
            # Calculate word count and extract text
            all_text = ' '.join(speaker_data['Transcription'].fillna('').astype(str))
            word_count = len(all_text.replace(' ', ''))  # Character count for Chinese text
            
            # Calculate words per second
            words_per_second = (word_count / speaking_duration) if speaking_duration > 0 else 0
            
            # Generate top 10 keywords using jieba tokenization for Mandarin
            # Tokenize using jieba and filter
            words = jieba.lcut(all_text)
            words = [w.strip() for w in words if len(w.strip()) >= 2]  # Keep words with 2+ characters
            
            # Use proper stopwords from dictionaries based on language
            stopwords = dictionaries.get_stopwords('MAN')  # Mandarin version
            filtered_words = [w for w in words if w not in stopwords]
            
            # Get top 10 most frequent words
            word_counter = Counter(filtered_words)
            top_10_words = [word for word, count in word_counter.most_common(10)]
            
            # Format as requested: "{'word1', 'word2', ..., 'word10'}"
            keywords_str = "{" + ", ".join([f"'{word}'" for word in top_10_words]) + "}"
            
            speaker_stats.append({
                'Speaker': speaker,
                'Speaking Duration (seconds)': speaking_duration,
                'Word Count': str(word_count),
                'Words per Second': f"{words_per_second:.2f}",
                'Top 10 Keywords': keywords_str
            })
    
    # Identify system recording speakers and add them to speaker_stats
    system_recording_speakers = set()
    if 'System_Audio_Recording' in updated_call_df.columns:
        system_df = updated_call_df[updated_call_df['System_Audio_Recording'] == 'YES']
        if not system_df.empty and 'Speaker Roles' in system_df.columns:
            system_recording_speakers = set(system_df['Speaker Roles'].dropna().unique())
            
            # Add system recording speakers to the speaker_stats
            for sys_speaker in system_recording_speakers:
                if sys_speaker and sys_speaker not in [stat['Speaker'] for stat in speaker_stats]:
                    speaker_stats.append({
                        'Role': 'System',
                        'Speaker': str(sys_speaker),
                        'Speaking Duration (seconds)': 'N/A',
                        'Word Count': 'N/A',
                        'Words per Second': 'N/A',
                        'Top 10 Keywords': 'N/A'
                    })
    
    # Sort speakers by word count to identify Sales and Customer (excluding system speakers)
    regular_speakers = [stat for stat in speaker_stats if stat.get('Role') != 'System']
    if regular_speakers:
        # Convert word count back to int for sorting
        for stat in regular_speakers:
            stat['_word_count_int'] = int(stat['Word Count'])
        
        regular_speakers.sort(key=lambda x: x['_word_count_int'], reverse=True)
        
        # Assign roles: highest word count = Sales, second highest = Customer
        if len(regular_speakers) >= 1:
            regular_speakers[0]['Role'] = 'Sales'
        if len(regular_speakers) >= 2:
            regular_speakers[1]['Role'] = 'Customer'
        
        # Remove the temporary sorting field and assign default roles
        for stat in regular_speakers:
            del stat['_word_count_int']
            if 'Role' not in stat:
                stat['Role'] = 'Other'
        
        # Ensure all speaker_stats have Role assigned
        for stat in speaker_stats:
            if 'Role' not in stat:
                stat['Role'] = 'Other'
    
    speaker_view_df = pd.DataFrame(speaker_stats)
    if not speaker_view_df.empty:
        # Reorder columns
        speaker_view_df = speaker_view_df[['Role', 'Speaker', 'Speaking Duration (seconds)', 'Word Count', 'Words per Second', 'Top 10 Keywords']]
    
    # ==========================================
    # Section C: Detailed Coverage List
    # ==========================================
    print("📋 Preparing Detailed Coverage List...")
    
    # Create a copy of results for modification
    detailed_coverage_df = results_df.copy()
    
    # Rename columns for better business understanding
    detailed_coverage_df = detailed_coverage_df.rename(columns={
        'Matched_Group': 'Matched_Group (Call Text)',
        'Best_Matching_Variation': 'Standard_Script'
    })
    
    # Enhance the Covered column with visual cues
    detailed_coverage_df['Covered'] = detailed_coverage_df['Covered'].apply(
        lambda x: f"✅ {x}" if x == 'Covered' else f"❌ {x}"
    )
    
    # Reorder columns to be more business-friendly
    business_friendly_columns = [
        'Required_Discussion_Point',
        'Standard_Script',
        'Covered',
        'Matched_Group (Call Text)',
        'Speaker',
        'Overlapping_Keywords',
        'Group_ID',
        'All_Variations_Count'
    ]
    
    # Add remaining columns (metric scores)
    remaining_columns = [col for col in detailed_coverage_df.columns if col not in business_friendly_columns]
    final_columns = business_friendly_columns + remaining_columns
    
    # Reorder the DataFrame
    detailed_coverage_df = detailed_coverage_df[final_columns]
    
    # Return the complete consolidated data for reference
    return {
        'kpis': kpis_df,
        'speaker_view': speaker_view_df,
        'detailed_coverage': detailed_coverage_df
    }

def write_executive_summary_content(ws, kpis_df, speaker_view_df, detailed_coverage_df):
    """Write Executive Summary content to a worksheet"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    import string
    
    current_row = 1
    
    # Define styles
    header_font = Font(bold=True, size=12)
    table_header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Section A: Overall Performance KPIs
    ws[f'A{current_row}'] = "A. Overall Performance KPIs"
    ws[f'A{current_row}'].font = header_font
    current_row += 2
    
    # Write KPIs data
    for idx, row in kpis_df.iterrows():
        ws[f'A{current_row}'] = row['Metric']
        ws[f'B{current_row}'] = row['Value']
        ws[f'A{current_row}'].font = table_header_font
        current_row += 1
    
    current_row += 2
    
    # Section B: Key Insights - Speaker View
    ws[f'A{current_row}'] = "B. Key Insights - Speaker View"
    ws[f'A{current_row}'].font = header_font
    current_row += 2
    
    if not speaker_view_df.empty:
        # Write headers
        for col_idx, col_name in enumerate(speaker_view_df.columns, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = col_name
            cell.font = table_header_font
            cell.fill = header_fill
            cell.border = border
        current_row += 1
        
        # Write data
        for idx, row in speaker_view_df.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.border = border
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No speaker data available"
        current_row += 1
    
    current_row += 2
    
    # Section C: Detailed Coverage List
    ws[f'A{current_row}'] = "C. Detailed Coverage List"
    ws[f'A{current_row}'].font = header_font
    current_row += 2
    
    if not detailed_coverage_df.empty:
        # Write headers
        for col_idx, col_name in enumerate(detailed_coverage_df.columns, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = col_name
            cell.font = table_header_font
            cell.fill = header_fill
            cell.border = border
        
        detail_header_row = current_row
        current_row += 1
        
        # Write data
        for idx, row in detailed_coverage_df.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.border = border
            current_row += 1
        
        # Enable AutoFilter on the detailed coverage table
        detail_end_col = len(detailed_coverage_df.columns)
        detail_end_row = current_row - 1
        detail_end_col_letter = chr(ord('A') + detail_end_col - 1)
        ws.auto_filter.ref = f"A{detail_header_row}:{detail_end_col_letter}{detail_end_row}"
    else:
        ws[f'A{current_row}'] = "No coverage data available"
        current_row += 1
    
    # Adjust column widths for readability
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for very long content
        ws.column_dimensions[column_letter].width = adjusted_width

# ==========================================================
# Main Routine - Four-Step Early Separation Strategy
# ==========================================================
# ==========================================================
# OLD MAIN FUNCTION REMOVED
# ==========================================================
# The previous main function that used global variables (SCRIPT_SHEET_NAME, 
# CALL_TEXT_FILE_PATH, etc.) has been removed and replaced with the modular 
# run_analysis function approach for batch processing compatibility.
# ==========================================================

# ==========================================================
# ==========================================================
def run_analysis(call_file_path, script_file_path, script_sheet_name, output_folder, call_type=None, language=None, include_system_audio=False):
    """
    Core analysis function implementing the four-step strategy:
    1. Early Separation: Split call_df into analysis_df and system_audio_df
    2. Independent Processing: Only process analysis_df for grouping
    3. Focused Analysis: check_coverage only analyzes human dialogue
    4. Late Merge & Reporting: Combine results for final reports
    
    Args:
        call_file_path: Path to the call text file
        script_file_path: Path to the script file
        script_sheet_name: Name of the script sheet
        output_folder: Folder to save output files
        call_type: Type of call ('SQCCB' or 'Sales Call'), if None will auto-detect
        language: Language code ('MAN', 'ENG'), if None will auto-detect
        include_system_audio: Whether to include system audio recordings in analysis (default: False)
        
    Returns:
        dict: Results dictionary with status, output_file, and coverage_rate
    """
    # Start tracing memory allocations and overall timer
    tracemalloc.start()
    overall_start = print_checkpoint(1, "Program startup, loading data files")
    
    # 🔍 通话类型检测逻辑（单一事实来源）
    if call_type is None:
        # Fallback: 从文件名检测（仅在直接调用时使用）
        filename = os.path.basename(call_file_path).upper()
        if "SQCCB" in filename:
            call_type = "SQCCB"
        else:
            call_type = "Sales Call"
            print(f"⚠️  Fallback detection - 通话类型: {call_type}")
    else:
        print(f"✅ 接收到的通话类型: {call_type}")
    
    print(f"📁 文件名: {os.path.basename(call_file_path)}")
    
    # Check if files exist
    if not os.path.exists(call_file_path):
        print(f"❌ Error: Call text file '{call_file_path}' not found.")
        return {'status': 'ERROR', 'error': f'Call file not found: {call_file_path}', 'output_file': None, 'coverage_rate': 0.0}
        
    if not os.path.exists(script_file_path):
        print(f"❌ Error: Script file '{script_file_path}' not found.")
        return {'status': 'ERROR', 'error': f'Script file not found: {script_file_path}', 'output_file': None, 'coverage_rate': 0.0}
    
    try:
        step1_start = time.time()
        
        # Load call text file (CSV format)
        print(f"📁 Loading call text from: {call_file_path}")
        if call_file_path.endswith('.csv'):
            call_df = pd.read_csv(call_file_path)
            print(f"   Loading CSV file directly")
        elif call_file_path.endswith('.xlsx') or call_file_path.endswith('.xls'):
            xl_call = pd.ExcelFile(call_file_path, engine='openpyxl')
            call_df = xl_call.parse(xl_call.sheet_names[0])  # Use first sheet
            print(f"   Using sheet: {xl_call.sheet_names[0]}")
        else:
            raise ValueError(f"Unsupported file format. Please use .csv, .xlsx, or .xls files.")
        
        # Load script file
        if script_file_path.endswith('.xlsx'):
            xl_script = pd.ExcelFile(script_file_path, engine='openpyxl')
            script_df = xl_script.parse(script_sheet_name)
        else:
            script_df = pd.read_csv(script_file_path)

        # Apply optional composite discussion-point key for configured sheets
        script_df = apply_composite_point_key_if_configured(script_df, script_sheet_name)
            
        print_checkpoint(2, f"Excel文件加载完成 ({len(call_df):,} 通话行, {len(script_df)} 脚本行)", step1_start)
    except Exception as e:
        print(f"❌ 读取Excel文件错误: {e}")
        return {'status': 'ERROR', 'error': str(e), 'output_file': None, 'coverage_rate': 0.0}

    # Initialize the checker
    step2_start = time.time()
    checker = SbertCallCoverageChecker(model_path=SBERT_MODEL_PATH)
    
    # Load precomputed script embeddings
    print("🔄 Loading precomputed script embeddings...")
    # Load precomputed script embeddings (SCRIPT_EMBEDDINGS_PATH imported at top)
    try:
        import pickle
        
        if os.path.exists(SCRIPT_EMBEDDINGS_PATH):
            with open(SCRIPT_EMBEDDINGS_PATH, 'rb') as f:
                checker.script_embeddings = pickle.load(f)
            print(f"✅ Loaded {len(checker.script_embeddings)} precomputed script embeddings")
        else:
            print(f"⚠️  Script embeddings file not found: {SCRIPT_EMBEDDINGS_PATH}")
            print("   Will fall back to on-demand encoding (slower)")
            checker.script_embeddings = {}
    except Exception as e:
        print(f"⚠️  Error loading script embeddings: {e}")
        print("   Will fall back to on-demand encoding (slower)")
        checker.script_embeddings = {}
    
    # 🔍 语言检测逻辑（单一事实来源）
    if language is None:
        # Fallback: 从文件名检测语言（仅在直接调用时使用）
        filename = os.path.basename(call_file_path).upper()
        if "_M" in filename:
            language = "MAN"
        elif "_E" in filename:
            language = "ENG"
        else:
            language = "MAN"  # 默认普通话
        print(f"⚠️  Fallback detection - 语言: {language}")
    else:
        print(f"✅ 接收到的语言: {language}")
    
    # 设置checker的语言
    checker.current_language = language

    # Normalize script embedding keys to clean-text keys for robust cache hit
    loaded_count, normalized_count, collisions = checker.normalize_script_embedding_keys()
    if loaded_count > 0:
        print(f"✅ Normalized script embedding keys: {loaded_count} -> {normalized_count} (collisions: {collisions})")
    
    # Load call-specific weights for all call types (unified processing)
    # This function is intelligent enough to handle both Sales Call and SQCCB based on script_sheet_name
    checker.load_call_specific_weights(call_file_path, script_df, script_sheet_name)
    
    # Check if the detected language is supported
    if checker.current_language not in ['MAN', 'ENG']:
        print(f"❌ Error: Language '{checker.current_language}' is not supported in Mandarin version.")
        print(f"Mandarin version only supports MAN (Mandarin) and ENG (English).")
        return {'status': 'ERROR', 'error': f'Unsupported language: {checker.current_language}', 'output_file': None, 'coverage_rate': 0.0}
    
    print(f"✅ Language validation passed: {checker.current_language}")
    print_checkpoint(3, "初始化分析器并验证语言支持", step2_start)
    
    # STEP 1 & 2: UNIFIED PROCESSING - Build grouped lines with call_type-aware system recording detection
    step3_start = time.time()
    print(f"\n📍 统一处理 - 使用 {call_type} 模式进行分组...")
    
    # Build grouped lines using the unified function that handles system recording detection internally
    grouped_lines, original_to_group, updated_call_df, system_recording_flags = checker.build_grouped_lines(call_df, call_type=call_type, include_system_audio=include_system_audio)
    
    # Create analysis_df and system_audio_df from updated_call_df (no duplicate calculation)
    analysis_df = updated_call_df[updated_call_df['System_Audio_Recording'] == 'NO'].copy()
    system_audio_df = updated_call_df[updated_call_df['System_Audio_Recording'] == 'YES'].copy()
    
    print(f"✅ 分组处理完成:")
    print(f"   - 创建了 {len(grouped_lines)} 个组")
    
    # Batch encode all grouped call texts for SBERT optimization
    print(f"🔄 批量编码通话文本...")
    
    # Build unique clean call texts (clean text is both SBERT input and embedding key)
    unique_preprocessed_call_texts = []
    seen_preprocessed_call_texts = set()
    
    for group in grouped_lines:
        if group.get('text'):
            preprocessed_text = checker.preprocess_text(group['text'], mode='comparison')
            if preprocessed_text and preprocessed_text not in seen_preprocessed_call_texts:
                unique_preprocessed_call_texts.append(preprocessed_text)
                seen_preprocessed_call_texts.add(preprocessed_text)
    
    if unique_preprocessed_call_texts:
        try:
            print(f"   编码 {len(unique_preprocessed_call_texts)} 个预处理后的通话组...")
            group_embeddings = checker.sbert_model.encode(unique_preprocessed_call_texts, show_progress_bar=True)
            
            # Create preprocessed to embedding mapping
            preprocessed_to_embedding = {}
            for text, embedding in zip(unique_preprocessed_call_texts, group_embeddings):
                preprocessed_to_embedding[text] = embedding
            
            # Final mapping: clean_text -> embedding (same key semantics as similarity lookup)
            group_text_to_embedding = preprocessed_to_embedding
            
            print(f"✅ 批量编码完成: {len(group_text_to_embedding)} 个向量")
        except Exception as e:
            print(f"⚠️  批量编码失败: {e}, 将使用逐个编码")
            group_text_to_embedding = {}
    else:
        print("⚠️  没有找到有效的通话文本进行编码")
        group_text_to_embedding = {}
    
    print_checkpoint(4, f"创建并保存分组通话数据 ({len(grouped_lines)} 个分组)", step3_start)
    
    # STEP 3: FOCUSED ANALYSIS - Coverage analysis on human dialogue only
    step5_start = time.time()
    print("\n📍 步骤3: 专注分析 - 对人类对话进行覆盖率分析...")
    
    # Check coverage using ONLY the human dialogue groups with precomputed embeddings
    results = checker.check_coverage(updated_call_df, script_df, grouped_lines, threshold=0.4, group_text_to_embedding=group_text_to_embedding)
    
    print(f"✅ 覆盖率分析完成:")
    print(f"   - 分析了 {len(results)} 个讨论点")
    print(f"   - 使用了 {len(grouped_lines)} 个人类对话组")
    print_checkpoint(5, f"完成覆盖分析 ({len(results)} 个要点)", step5_start)
    
    # STEP 4: LATE MERGE & REPORTING - Combine results for final reports
    step6_start = time.time()
    print("\n📍 步骤4: 后期合并与报告 - 创建综合报告...")
    
    # Create call text analysis with proper merging
    call_text_analysis = checker.create_call_text_analysis_view_with_separation(
        grouped_lines, system_audio_df, script_df, threshold=0.4, group_text_to_embedding=group_text_to_embedding
    )
    
    # Create sentence level output
    sentence_level_output = checker.create_sentence_level_output(updated_call_df, original_to_group)
    print_checkpoint(6, f"生成最终分析视图", step6_start)
    
    # Generate Reports with Executive Summary
    step7_start = time.time()
    print("\n📊 生成包含Executive Summary的最终报告...")
    
    # Generate unique output filename based on input filename and save to output_folder
    input_filename = os.path.basename(call_file_path)
    if input_filename.endswith('.csv'):
        # Remove .wav if present, then replace .csv with .xlsx
        if '.wav' in input_filename:
            output_filename = input_filename.replace('.wav.csv', '.xlsx')
        else:
            output_filename = input_filename.replace('.csv', '.xlsx')
    elif input_filename.endswith('.xlsx') or input_filename.endswith('.xls'):
        # Remove .wav if present, then add _result
        if '.wav' in input_filename:
            output_filename = input_filename.replace('.wav.xlsx', '.xlsx').replace('.wav.xls', '.xlsx')
        else:
            output_filename = input_filename.replace('.xlsx', '_result.xlsx').replace('.xls', '_result.xlsx')
    else:
        # Remove .wav if present, then add _result
        if '.wav' in input_filename:
            output_filename = input_filename.replace('.wav', '') + '_result.xlsx'
        else:
            output_filename = f"{input_filename}_result.xlsx"
    
    # Create full output path
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, output_filename)
    
    print(f"📝 Output file will be: {output_file}")
    try:
        # Generate the Executive Summary data first
        summary_data = generate_executive_summary_report(
            results, updated_call_df, script_df, grouped_lines, None  # Don't save to separate file
        )
        
        # Save to the main output file with Executive Summary as first sheet
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # First, create Executive Summary sheet manually
            wb = writer.book
            ws = wb.create_sheet("Executive Summary", 0)
            write_executive_summary_content(ws, summary_data['kpis'], summary_data['speaker_view'], summary_data['detailed_coverage'])
            
            # Then write other sheets
            call_text_analysis.to_excel(writer, sheet_name='Coverage Analysis', index=False)
            sentence_level_output.to_excel(writer, sheet_name='Original Call Text', index=False)
            
            # Remove the default 'Sheet' if it exists
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
        
        # Verify file creation
        if os.path.exists(output_file):
            file_size = os.path.getsize(output_file)
            print(f"✅ Excel文件成功创建: {output_file} ({file_size:,} bytes)")
        else:
            print(f"❌ Excel文件未创建: {output_file}")
            return {'status': 'ERROR', 'error': 'File creation failed', 'output_file': None, 'coverage_rate': 0.0}
            
    except Exception as e:
        print(f"❌ 生成报告错误: {e}")
        print("尝试保存备份CSV文件...")
        backup_folder = os.path.join(output_folder, 'backup')
        os.makedirs(backup_folder, exist_ok=True)
        results.to_csv(os.path.join(backup_folder, f'coverage_analysis_backup_{os.path.splitext(output_filename)[0]}.csv'), index=False, encoding='utf-8-sig')
        call_text_analysis.to_csv(os.path.join(backup_folder, f'call_text_analysis_backup_{os.path.splitext(output_filename)[0]}.csv'), index=False, encoding='utf-8-sig')
        sentence_level_output.to_csv(os.path.join(backup_folder, f'sentence_level_analysis_backup_{os.path.splitext(output_filename)[0]}.csv'), index=False, encoding='utf-8-sig')
        return {'status': 'ERROR', 'error': str(e), 'output_file': None, 'coverage_rate': 0.0}
    
    # Create grouped call DataFrame for reference
    grouped_call_df = pd.DataFrame(grouped_lines)
    reference_file = os.path.join(output_folder, f"grouped_call_data_{os.path.splitext(output_filename)[0]}.xlsx")
    grouped_call_df.to_excel(reference_file, index=False)
    
    print_checkpoint(7, f"生成包含执行摘要的报告 ({output_file})", step7_start)
    
    # Calculate final metrics
    total_time = time.time() - overall_start
    covered_count = len(results[results['Covered'] == 'Covered'])
    total_points = len(results)
    coverage_rate = (covered_count / total_points * 100) if total_points > 0 else 0.0
    
    print(f"\n🎉 四步早期分离策略分析完成! 总时间: {total_time:.1f}s")
    print(f"   • STEP 1: Early Separation - System recording detection")
    print(f"   • STEP 2: Independent Processing - Human dialogue grouping")
    print(f"   • STEP 3: Focused Analysis - Coverage analysis on human dialogue only")
    print(f"   • STEP 4: Late Merge & Reporting - Comprehensive report generation")
    print(f"   📋 Files created:")
    print(f"   • {output_file}: Executive Summary with late merge strategy")
    print(f"   • {reference_file}: Pure human dialogue groups")
    
    print(f"\n📊 Coverage Summary:")
    print(f"   • Coverage Rate: {coverage_rate:.1f}% ({covered_count}/{total_points} points)")
    print(f"   • Total groups created: {len(grouped_lines)}")
    print(f"   • Call type: {call_type}")

    print(f"\n📊 Output Files Created:")
    print(f"  • {output_file} (3 sheets with Executive Summary):")
    print(f"    - Executive Summary: Consolidated business-friendly report with KPIs, Speaker View & Detailed Coverage")
    print(f"    - Coverage Analysis: {len(call_text_analysis)} call text segments analysis")
    print(f"    - Original Call Text: {len(sentence_level_output)} sentence-level analysis")
    print(f"  • {reference_file}: {len(grouped_call_df)} grouped call data reference")
    print(f"\n🎉 Executive Summary Features:")
    print(f"  • 📈 Overall Performance KPIs: Total duration, coverage rate, risk points")
    print(f"  • 🎤 Speaker View: Role identification (Sales/Customer/System), duration, keywords")
    print(f"  • 📋 Detailed Coverage List: Business-friendly format with ✅/❌ indicators")
    print(f"  • 💼 Professional Excel formatting: Bold headers, borders, auto-filter")
    print(f"  • 🔄 Dynamic filename: {input_filename} → {output_filename}")

    # Final checkpoint and cleanup
    overall_time = print_checkpoint("FINAL", "All processing completed", overall_start)
    
    # Print memory usage
    current, peak = tracemalloc.get_traced_memory()
    print(f"\n💾 Memory usage summary:")
    print(f"  - Current memory usage: {current / 10**6:.2f}MB")
    print(f"  - Peak memory usage: {peak / 10**6:.2f}MB")

    # Stop tracing memory allocations
    tracemalloc.stop()
    
    print(f"\n🎉 Program execution completed! Total time: {overall_time - overall_start:.2f}s")
    
    # Extract speaker word counts from summary_data
    sales_word_count = 0
    customer_word_count = 0
    
    if summary_data and 'speaker_view' in summary_data:
        speaker_view_df = summary_data['speaker_view']
        if not speaker_view_df.empty:
            for _, row in speaker_view_df.iterrows():
                if row.get('Role') == 'Sales' and pd.notna(row.get('Word Count')) and str(row.get('Word Count')).isdigit():
                    sales_word_count = int(row['Word Count'])
                elif row.get('Role') == 'Customer' and pd.notna(row.get('Word Count')) and str(row.get('Word Count')).isdigit():
                    customer_word_count = int(row['Word Count'])
    
    # Convert total_call_duration to seconds (numeric)
    total_call_duration_seconds = 0
    if summary_data and 'kpis' in summary_data:
        kpis_df = summary_data['kpis']
        duration_row = kpis_df[kpis_df['Metric'] == 'Total Call Duration']
        if not duration_row.empty:
            duration_value = duration_row.iloc[0]['Value']
            if pd.notna(duration_value) and str(duration_value) != 'N/A':
                try:
                    # Parse time string to seconds
                    duration_str = str(duration_value)
                    if ':' in duration_str:
                        parts = duration_str.split(':')
                        if len(parts) == 2:  # MM:SS
                            total_call_duration_seconds = int(parts[0]) * 60 + int(parts[1])
                        elif len(parts) == 3:  # HH:MM:SS
                            total_call_duration_seconds = int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
                    else:
                        # Direct numeric value
                        total_call_duration_seconds = float(duration_str)
                except (ValueError, TypeError):
                    total_call_duration_seconds = 0
    
    # Return enhanced results dictionary
    return {
        'status': 'SUCCESS',
        'output_file': output_file,
        'coverage_rate': coverage_rate,
        'total_points': total_points,
        'covered_points': covered_count,
        'total_call_duration': total_call_duration_seconds,
        'sales_word_count': sales_word_count,
        'customer_word_count': customer_word_count,
        'processing_time': total_time,
        'call_type': call_type,
        'language': checker.current_language
    }


def main():
    """
    Main function for standalone testing using default configuration.
    Note: When used standalone, you need to modify the paths below or use run_batch_analysis.py
    """
    # Default paths for standalone testing - modify these as needed
    default_call_file = "call_text_sample_M.wav.csv"
    default_script_file = "Scripts.xlsx" 
    default_script_sheet = "Script"
    default_output_folder = "."
    
    print("⚠️  Running in standalone mode with default paths.")
    print("For batch processing, use run_batch_analysis.py instead.")
    
    return run_analysis(default_call_file, default_script_file, default_script_sheet, default_output_folder)


if __name__ == "__main__":
    main()
