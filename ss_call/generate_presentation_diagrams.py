#!/usr/bin/env python3
"""
Generate Presentation_Diagrams.xlsx with English Process Flow and supporting sheets.

Key points:
- No merged cells are used (avoids openpyxl 'MergedCell value is read-only' issues).
- Adds:
  - 04_Dynamic_Analytics: high-level dynamic analytics overview (English).
  - 05_Artifacts: outputs you can expect from batch analysis (English).
  - 08_Process_Flow: one-time preparation + per-review end-to-end steps, with explicit Cantonese/Mandarin divergence (English).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _apply_header_style(ws, header_row=1, fill_color="FFD966"):
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_body_alignment(ws, start_row=2, end_row=None, start_col=1, end_col=None, wrap=True, valign="top"):
    if end_row is None:
        end_row = ws.max_row
    if end_col is None:
        end_col = ws.max_column
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).alignment = Alignment(vertical=valign, wrap_text=wrap)


def _set_column_widths(ws, widths):
    # widths: dict like {"A": 18, "B": 24}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _apply_table_borders(ws, start_row=1, end_row=None, start_col=1, end_col=None):
    if end_row is None:
        end_row = ws.max_row
    if end_col is None:
        end_col = ws.max_column
    thin = Side(style="thin", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = border


def add_dynamic_analytics_sheet(wb):
    """
    Create an English high-level Dynamic Analytics sheet without merged cells.
    """
    ws = wb.create_sheet(title="04_Dynamic_Analytics")
    _set_column_widths(ws, {"A": 28, "B": 48, "C": 28})

    headers = ["Component", "Purpose", "Notes"]
    ws.append(headers)

    rows = [
        [
            "Dynamic Script Analysis",
            "Parse product/language script variants, tokenize, generate embeddings for canonical phrases/synonyms.",
            "Backed by multilingual SBERT; cached embeddings for reuse; avoids recomputation."
        ],
        [
            "Coverage Checker",
            "Compare ASR transcript segments against script variants; compute similarity; capture evidence snippets.",
            "Language-aware normalization, synonyms expansion, negation handling; dual-check at segment and topic levels."
        ],
        [
            "System Audio Handling",
            "Separate system prompts early; merge later with weights to ensure fair coverage while preserving human dialogue focus.",
            "Configurable include/exclude flags; logs evidence provenance."
        ],
        [
            "Topic Weighting",
            "Apply dynamic weights per topic/test-point to reflect compliance importance and product context.",
            "Lazy-loaded weights (CSV.gz); updated via team feedback; single source of truth with fallback."
        ],
        [
            "Batch Orchestration",
            "Queue calls by product/language; robust per-file execution; aggregate outputs to human-readable reports.",
            "Outlier flagging via IQR; retry/fallback; structured artifacts for traceability."
        ],
    ]
    for row in rows:
        ws.append(row)

    _apply_header_style(ws)
    _apply_body_alignment(ws)
    _apply_table_borders(ws, start_row=1)


def add_artifacts_sheet(wb):
    """
    Describe produced artifacts (CSV/XLSX/logs) for analysis runs.
    """
    ws = wb.create_sheet(title="05_Artifacts")
    _set_column_widths(ws, {"A": 24, "B": 36, "C": 52, "D": 24})

    headers = ["Artifact", "Filename Pattern", "Content", "Notes"]
    ws.append(headers)

    rows = [
        [
            "Sentence-level analysis",
            "sentence_level_analysis_<call_or_batch>.csv",
            "Per-segment normalized text, topic hits, similarity scores, negation flags, evidence spans.",
            "Language-aware processing; useful for deep dives."
        ],
        [
            "Coverage analysis",
            "coverage_analysis_<call_or_batch>.csv",
            "Topic/test-point coverage status (covered/partial/missing), evidence references, weights applied.",
            "Readable summary for compliance reviewers."
        ],
        [
            "Grouped call data",
            "grouped_call_data_<batch>.xlsx",
            "Aggregated metrics per product/language/date/channel/agent; outlier flags via IQR.",
            "Good for QA and governance dashboards."
        ],
        [
            "Logs & backups",
            "backup/*.csv, run logs",
            "Execution traces, intermediate outputs, fallback artifacts.",
            "Supports reproducibility and incident analysis."
        ],
    ]
    for row in rows:
        ws.append(row)

    _apply_header_style(ws)
    _apply_body_alignment(ws)
    _apply_table_borders(ws, start_row=1)


def add_process_flow_sheet(wb):
    """
    English Process Flow sheet (no merged cells).
    Includes:
    - One-time preparation
    - Per-review end-to-end pipeline
    Explicitly distinguishes Cantonese vs. Mandarin in similarity, metrics, and script resource reuse.
    """
    ws = wb.create_sheet(title="08_Process_Flow")
    _set_column_widths(ws, {"A": 18, "B": 28, "C": 64, "D": 36, "E": 36})

    headers = ["Phase", "Step", "Details", "Cantonese (ZH-YUE)", "Mandarin (ZH-CMN)"]
    ws.append(headers)

    rows = [
        # One-time preparation
        ["One-time preparation", "Standardize scripts",
         "Align scripts by product and language; canonical phrasing and versioning.", "Shared", "Shared"],
        ["One-time preparation", "Dictionaries per test points",
         "Build compliance dictionaries per product; maintain separate term sets per language.", "Dedicated Cantonese dictionary", "Dedicated Mandarin dictionary"],
        ["One-time preparation", "Dynamic term weights (CSV.gz)",
         "Generate topic-specific weights; single source of truth with fallback; lazy-loaded during analysis.", "Shared source/fallback", "Shared source/fallback"],
        ["One-time preparation", "SBERT resources",
         "Prepare multilingual SBERT model and optional precomputed embeddings for script phrases.", "Model available", "Model + precomputed script embeddings prepared"],

        # Per-review E2E
        ["Per-review E2E", "Parse call naming",
         "Extract product, language, date, channel, agent ID from filename.", "Shared", "Shared"],
        ["Per-review E2E", "Pre-processing",
         "Load ASR transcripts; de-duplicate; clean text; merge short sentences.", "Shared", "Shared"],
        ["Per-review E2E", "System audio handling",
         "Early separate system prompts; later merge with weights if enabled.", "Shared; controlled by config", "Shared; controlled by config"],
        ["Per-review E2E", "Language branch start",
         "Pipeline diverges by language-specific resources and tokenization.", "Cantonese pipeline starts", "Mandarin pipeline starts"],

        # Script resource reuse differences
        ["Per-review E2E", "Script resource reuse",
         "Initialize language-specific resources for scoring and coverage.", "No precomputed script embeddings required (lexical focus)", "Reuse precomputed script embeddings (SBERT-based)"],

        # Tokenization and normalization
        ["Per-review E2E", "Normalize & tokenize",
         "Apply language-appropriate normalization and tokenization rules.", "Cantonese tokenizer/rules", "Mandarin tokenizer/rules"],

        # Synonyms and negation handling
        ["Per-review E2E", "Synonyms & phrase expansion",
         "Expand phrases via language dictionaries and product terms.", "Cantonese dictionary", "Mandarin dictionary"],
        ["Per-review E2E", "Negation detection",
         "Detect negation bigrams and polarity flips in local contexts.", "Cantonese negation rules", "Mandarin negation rules"],

        # Similarity engine and metric weighting differences
        ["Per-review E2E", "Similarity engine & metrics",
         "Compute similarity and apply topic/test-point weights (centrally managed; supports gradual adjustments).",
         "Primary: TF-IDF + cosine; Complementary: ROUGE-L and lexical metrics; weights centrally managed",
         "Primary: SBERT semantic similarity; Complementary: word/sentence-level indicators; weights centrally managed"],

        # Coverage and aggregation
        ["Per-review E2E", "Coverage checks",
         "Validate coverage across script variations; capture evidence snippets.", "Cantonese variants", "Mandarin variants (reusing embeddings)"],
        ["Per-review E2E", "Aggregation & summary",
         "Aggregate hits by topic; produce human-readable coverage summary.", "Shared", "Shared"],

        # Governance and outputs
        ["Per-review E2E", "Outlier flags",
         "Flag duration/metric outliers via IQR; mark calls needing review.", "Shared", "Shared"],
        ["Per-review E2E", "Artifacts export",
         "Export CSV/XLSX: sentence-level, coverage, grouped call data; logs.", "Shared", "Shared"],
        ["Per-review E2E", "Team feedback",
         "Capture reviewer comments; update dictionaries/weights; iterate.", "Cantonese updates", "Mandarin updates"],
    ]
    for row in rows:
        ws.append(row)

    _apply_header_style(ws)
    _apply_body_alignment(ws)
    _apply_table_borders(ws, start_row=1)


def main():
    wb = Workbook()
    # Remove the default sheet to control sheet order
    default_ws = wb.active
    wb.remove(default_ws)

    # Add sheets
    add_dynamic_analytics_sheet(wb)
    add_artifacts_sheet(wb)
    add_process_flow_sheet(wb)

    # Save to file in current directory
    output_path = "Presentation_Diagrams.xlsx"
    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
