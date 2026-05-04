"""
============================================================
Advanced Call Coverage Checker (Cantonese Speech-to-Text Analysis)
============================================================

Background & Purpose
--------------------
This script analyzes Cantonese call transcripts (from speech-to-text) and automatically checks whether all required discussion points are covered. It uses advanced NLP techniques including TF-IDF + Cosine Similarity for semantic matching, tailored for financial/investment product sales calls where regulatory compliance is critical.

Key Features
----------------
1. **Business-Weighted TF-IDF + Cosine Similarity:** Industry-standard semantic similarity with custom business term weighting
2. **Robust Speech-to-Text Error Handling:** Fuzzy string matching handles transcription errors and typos
3. **Cantonese Synonym Intelligence:** Custom synonym expansion for colloquial Cantonese business terms
4. **Smart Speaker-Based Grouping:** Groups consecutive sentences by speaker with 150-character limit per group
5. **Streamlined 3-Metric System:** Eliminates redundancy while maintaining accuracy
6. **Performance Optimizations:** Cached TF-IDF vectorizer, sparse matrix operations, similarity caching, early exit optimizations (uses fresh dictionaries.py weights)
7. **Memory-Efficient Processing:** Dual-algorithm approach with automatic switching based on dataset size

Similarity Metrics (Streamlined)
----------------
1. **TF-IDF + Cosine Similarity (50%):** Semantic similarity with business term importance weighting
2. **Expanded Token Overlap (35%):** Synonym-aware matching for Cantonese business terms
3. **Fuzzy String Matching (15%):** Character-level similarity for speech-to-text error tolerance

Inputs & Outputs
----------------
- **Input:**
    - Excel file with 'Call_Text' sheet (File, Time, speaker_role, Text) and 'Script' sheet (Required_Discussion_Point, Standard_Script)
- **Output:**
    - Detailed coverage results with similarity metrics and matched speaker groups
    - Grouped call data for reference and debugging

Script Structure
----------------
1. **Class: AdvancedCallCoverageChecker**
    - Business term dictionaries with importance weights
    - Advanced similarity calculation (TF-IDF + Cosine, Synonyms, Fuzzy)
    - Speaker-based grouping and coverage analysis
2. **Main Routine**
    - Data loading and preprocessing
    - Coverage analysis with detailed metrics
    - Results export and summary

============================================================
"""

# =============================
# Imports & Dependencies
# =============================
# Standard and third-party libraries for data processing, text analysis, and similarity computation.
import pandas as pd
import pycantonese as pc
import re
import os
import sys
from difflib import SequenceMatcher
from collections import Counter
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import tracemalloc
import time

# Ensure proper encoding for Windows
if sys.platform.startswith('win'):
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.detach())
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.detach())

# Import dictionaries
import dictionaries
from config import (
    COMPOSITE_POINT_KEY_SEPARATOR,
    COMPOSITE_POINT_NOTE_COLUMN,
    should_use_composite_point_key,
)

# =============================
# USER CONFIGURATION SECTION
# =============================
"""
CONFIGURE YOUR DISCUSSION POINTS FOR ENHANCED SCORING HERE:
============================================================

Instructions:
1. Add discussion point names that you want to enhance with date/numeric pattern recognition
2. Set 'date_boost': score boost (0.05-0.25) for points that should get bonus for date patterns
3. Set 'numeric_boost': score boost (0.05-0.25) for points that should get bonus for numeric patterns
4. Set both to 0.0 if you don't want that type of enhancement

Examples:
- For points about dates/deadlines: set date_boost > 0, numeric_boost = 0
- For points about prices/amounts: set numeric_boost > 0, date_boost = 0
- For mixed points: set both > 0
"""

# MODIFY THIS DICTIONARY TO CONFIGURE YOUR DISCUSSION POINTS:
USER_ENHANCED_DISCUSSION_POINTS = {
    # Example entries - modify these as needed:
    'Price': {'date_boost': 0.0, 'numeric_boost': 0.20},
    'RPQ': {'date_boost': 0.15, 'numeric_boost': 0.0},
    
    # Add your discussion points here:
    'acknowledge SFP report': {'date_boost': 0.15, 'numeric_boost': 0.0},
    '交易价格确认': {'date_boost': 0.0, 'numeric_boost': 0.18},
    '交易数量确认': {'date_boost': 0.0, 'numeric_boost': 0.20},
    '產品信息確認': {'date_boost': 0.0, 'numeric_boost': 0.12},
    '重要事項聲明': {'date_boost': 0.12, 'numeric_boost': 0.0},
    '財務狀況評估': {'date_boost': 0.0, 'numeric_boost': 0.15},
    
    # Template for adding more:
    # 'Your Discussion Point Name': {'date_boost': 0.1, 'numeric_boost': 0.15},
}

# =============================
# INPUT FILE CONFIGURATION
# =============================
"""
CONFIGURE YOUR INPUT FILE PATHS HERE:
====================================

Instructions:
1. Set CALL_TEXT_FILE_PATH to the path of your call transcript file
2. Set SCRIPT_FILE_PATH to the path of your script/required points file
3. Use forward slashes (/) or raw strings for Windows paths
4. Can be relative paths (from script location) or absolute paths

File Format Requirements:
- Call text file: Must contain columns 'Speaker Roles', 'Transcription', 'Segment Start Time', 'Segment End Time'
- Script file: Must contain columns 'Required_Discussion_Point', 'Standard_Script'
"""

# ==========================================================
# DEPRECATED CONFIGURATION SECTION
# ==========================================================
# This section is no longer used for batch processing.
# Configuration is now handled by run_batch_analysis.py and config.py
# For standalone testing, modify the default values in main() function.
# ==========================================================

# Progress tracking
def print_checkpoint(step_num, description, start_time=None):
    """Print progress checkpoint with timestamp and elapsed time"""
    current_time = time.time()
    timestamp = time.strftime('%H:%M:%S')
    
    if start_time is not None:
        elapsed = current_time - start_time
        print(f"✅ CHECKPOINT {step_num}: {description} (Completed at: {timestamp}, Duration: {elapsed:.2f}s)")
    else:
        print(f"🚀 CHECKPOINT {step_num}: {description} (Started at: {timestamp})")
    
    return current_time

# ==========================================================
# Class: AdvancedCallCoverageChecker
# ----------------------------------------------------------
# Encapsulates all logic for preprocessing, synonym expansion,
# similarity calculation, and coverage checking.
# ==========================================================
class AdvancedCallCoverageChecker:
    def __init__(self):
        """
        Initialization: Load dictionaries and configurations from dictionaries.py.
        These are tailored for Cantonese financial calls and can be extended as needed.
        Now supports multi-product dynamic weights with language-specific configurations.
        """
        # Load all dictionaries from the centralized dictionaries file
        self.cantonese_synonyms = dictionaries.cantonese_synonyms
        self.error_patterns = dictionaries.error_patterns
        self.important_keywords = dictionaries.important_keywords
        # Language-specific stopwords will be loaded in load_call_specific_weights
        self.stopwords = None  # Will be set based on detected language
        
        # Multi-product weight management
        self.current_language = None
        self.current_product = None
        self.current_weights = {}  # Will be loaded by load_call_specific_weights
        
        # No cache - always use fresh weights from dictionaries.py
        # Call-level caches for efficiency
        self.call_char_ngram_freq = None  # dict: ngram -> freq for current call
        self.high_freq_threshold = 6
        
        # Performance optimization caches (not related to dictionaries cache)
        self._tfidf_vectorizer = None  # Cached TF-IDF vectorizer
        self._preprocessed_texts_cache = {}  # Cache for preprocessed texts
        self._similarity_cache = {}  # Cache for similarity calculations
        self._reverse_synonym_map = None  # Pre-computed reverse synonym mapping

    # ------------------------------------------------------
    # Text Preprocessing
    # ------------------------------------------------------
    def preprocess_text(self, text, mode='comparison', text_type='call'):
        """
        统一文本预处理函数，支持多种模式
        
        Args:
            text: 输入文本
            mode: 处理模式
                - 'comparison': 用于比较分析（去除英文，script和call统一处理）
                - 'display': 用于显示（保留英文）
            text_type: 文本类型 ('call' 或 'script')，主要用于缓存区分
                
        Returns:
            处理后的文本
        """
        text_str = str(text)
        # 创建缓存键，包含模式和文本类型
        cache_key = f"{mode}_{text_type}_{text_str}"
        
        # 检查缓存
        if cache_key in self._preprocessed_texts_cache:
            return self._preprocessed_texts_cache[cache_key]
        
        if mode == 'comparison':
            # 比较模式：去除英文字母（script和call统一处理，确保比较一致性）
            processed = re.sub(r'[^\u4e00-\u9fff,.。，。%()（）]+', '', text_str)
        elif mode == 'display':
            # 显示模式：保留英文字母（用于sentence level analysis等显示用途）
            processed = re.sub(r'[^\u4e00-\u9fffa-zA-Z,.。，。%()（）]+', '', text_str)
        else:
            # 默认使用比较模式
            processed = re.sub(r'[^\u4e00-\u9fff,.。，。%()（）]+', '', text_str)
        
        # 标准化常见的语音转文字错误
        for correct, variations in self.error_patterns.items():
            for variation in variations:
                processed = processed.replace(variation, correct)
        
        result = processed
        
        # 缓存结果（限制缓存大小以防内存溢出）
        if len(self._preprocessed_texts_cache) < 10000:
            self._preprocessed_texts_cache[cache_key] = result
            
        return result

    # 保持向后兼容的wrapper函数
    def preprocess_script_text(self, text):
        """向后兼容函数：用于script文本预处理"""
        return self.preprocess_text(text, mode='comparison', text_type='script')
    
    def preprocess_call_text(self, text):
        """向后兼容函数：用于call文本预处理"""
        return self.preprocess_text(text, mode='comparison', text_type='call')

    # ------------------------------------------------------
    # Synonym Expansion
    # ------------------------------------------------------
    def _build_reverse_synonym_map(self):
        """
        Build reverse synonym mapping for O(1) lookup performance.
        Maps each synonym to its canonical key and all related synonyms.
        """
        if self._reverse_synonym_map is not None:
            return self._reverse_synonym_map
            
        reverse_map = {}
        for key, synonyms in self.cantonese_synonyms.items():
            # Map the key to itself and all synonyms
            all_variants = {key} | set(synonyms)
            reverse_map[key] = all_variants
            
            # Map each synonym to the same set
            for synonym in synonyms:
                reverse_map[synonym] = all_variants
                
        self._reverse_synonym_map = reverse_map
        return reverse_map
    
    def expand_keywords(self, text):
        """
        Expand text with Cantonese synonyms for robust matching (optimized).
        - Tokenizes text using pycantonese.
        - Uses pre-computed reverse mapping for O(1) synonym lookup.
        """
        reverse_map = self._build_reverse_synonym_map()
        expanded_tokens = set()
        tokens = pc.segment(text)
        
        for token in tokens:
            expanded_tokens.add(token)
            # Use reverse mapping for fast lookup
            if token in reverse_map:
                expanded_tokens.update(reverse_map[token])
                
        return expanded_tokens
    
    def tokenize_text(self, text):
        """
        Tokenize text using pycantonese and filter out stopwords.
        Returns business-relevant tokens only.
        """
        tokens = pc.segment(text)
        # Filter out stopwords and keep only business-relevant terms
        business_tokens = [token for token in tokens if token not in self.stopwords and len(token) > 1]
        return business_tokens

    # ------------------------------------------------------
    # Multi-Product Weight Management
    # ------------------------------------------------------
    def detect_language_from_filename(self, file_path):
        """
        Detect language from CSV filename
        Examples:
        - "xxxxxx_M.wav.csv" -> "MAN" (Mandarin)
        - "xxxxxx_C.wav.csv" -> "CAN" (Cantonese)  
        - "xxxxxx_E.wav.csv" -> "ENG" (English)
        """
        filename = os.path.basename(file_path)
        # Remove .csv extension
        name_without_csv = filename.replace('.csv', '')
        # Remove .wav if present
        name_without_wav = name_without_csv.replace('.wav', '')
        
        # Get the last character after underscore
        if name_without_wav.endswith('_M'):
            return "MAN"
        elif name_without_wav.endswith('_C'):
            return "CAN"
        elif name_without_wav.endswith('_E'):
            return "ENG"
        else:
            return None

    def detect_product_type_from_script(self, script_df, script_sheet_name=None):
        """
        Detect product type from script DataFrame by extracting from script_sheet_name.
        Extracts product name from string before the underscore in script_sheet_name.
        Examples:
        - "Bond_MAN" -> "Bond"
        - "SID CPI3_CAN" -> "SID CPI3"
        """
        if script_sheet_name and '_' in script_sheet_name:
            # Extract everything before the underscore
            product_name = script_sheet_name.split('_')[0].strip()
            print(f"🔍 Detected product type from sheet name '{script_sheet_name}': {product_name}")
            return product_name
        elif script_sheet_name:
            print(f"⚠️  Sheet name '{script_sheet_name}' does not contain underscore, using as-is: {script_sheet_name}")
            return script_sheet_name
        else:
            print(f"⚠️  No script_sheet_name provided, cannot detect product type")
            return "Unknown"

    def load_call_specific_weights(self, call_file_path, script_df, script_sheet_name=None):
        """
        Load specific weights and language-specific configurations based on call file and script
        """
        print("🔍 Loading call-specific weights and language configurations...")
        
        # 1. Detect language from filename
        self.current_language = self.detect_language_from_filename(call_file_path)
        
        # 2. Load language-specific stopwords
        if self.current_language:
            try:
                self.stopwords = dictionaries.get_stopwords(self.current_language)
                print(f"✅ Loaded {self.current_language} stopwords: {len(self.stopwords)} words")
            except Exception as e:
                print(f"⚠️  Error loading {self.current_language} stopwords: {e}")
                print(f"⚠️  Using default Cantonese stopwords")
                self.stopwords = dictionaries.cantonese_stopwords
        else:
            # Default to Cantonese if language detection fails
            self.stopwords = dictionaries.cantonese_stopwords
        
        # 3. Detect product type from script
        self.current_product = self.detect_product_type_from_script(script_df, script_sheet_name)
        
        # 4. Load corresponding weights using unified interface
        # Trust dictionaries.get_product_weights to handle all fallback logic internally
        if self.current_language and self.current_product:
            # Call get_product_weights - it will handle fallback internally if needed
            self.current_weights = dictionaries.get_product_weights(
                self.current_language, self.current_product
            )
            print(f"✅ Loaded weights for {self.current_language}:{self.current_product}")
            print(f"📊 Using {len(self.current_weights)} term weights")
        else:
            # Detection failed - use general weights directly
            if not self.current_language:
                print(f"⚠️  Could not detect language from filename: {call_file_path}")
            if not self.current_product:
                print(f"⚠️  Could not detect product type from script")
            print(f"⚠️  Using general term_importance weights")
            self.current_weights = getattr(dictionaries, 'term_importance', {})

    # ------------------------------------------------------
    # Salescall System Audio Recording Detection
    # ------------------------------------------------------
    def detect_system_audio_recordings(self, call_df):
        """
        Salescall Detection: Identify system audio recordings in sales calls.
        
        Logic:
        - Excludes speaker_1 and speaker_2 (main participants)
        - For remaining speakers, finds consecutive segments with:
          * 10+ consecutive rows by same speaker
          * Average sentence length >= 20 characters
        - Once a speaker is identified as system audio source, ALL their lines
          in the entire call will be marked as system recordings
        
        Args:
            call_df: DataFrame with call data
            
        Returns:
            pandas.Series: Boolean series indicating system audio recording rows
        """
        print("🔍 Salescall: Starting system audio recording detection...")
        
        speaker_col = 'Speaker Roles'
        text_col = 'Transcription'
        
        if speaker_col not in call_df.columns or text_col not in call_df.columns:
            print("⚠️  Salescall: Required columns not found, skipping recording detection")
            return pd.Series([False] * len(call_df), index=call_df.index)
        
        # Initialize recording detection array
        is_recording = pd.Series([False] * len(call_df), index=call_df.index)
        
        # Get all speakers except main participants (SPEAKER_1 and SPEAKER_2 are sales and customer)
        # Normalize speaker names to uppercase for consistent comparison
        all_speakers = call_df[speaker_col].dropna().unique()
        normalized_speakers = [str(s).upper() for s in all_speakers]
        main_speakers = ['SPEAKER_1', 'SPEAKER_2']
        system_speakers = [s for s in all_speakers if str(s).upper() not in main_speakers]
        
        print(f"📊 Salescall: Found {len(system_speakers)} potential system speakers: {system_speakers}")
        
        # Create a set to track confirmed recording speakers
        recording_speakers = set()
        
        for speaker in system_speakers:
            speaker_rows = call_df[call_df[speaker_col] == speaker]
            if len(speaker_rows) < 10:  # Need at least 10 rows
                continue
            
            # Find consecutive segments for this speaker
            speaker_indices = speaker_rows.index.tolist()
            consecutive_segments = self._find_consecutive_segments(speaker_indices)
            
            for segment_indices in consecutive_segments:
                if len(segment_indices) >= 10:  # At least 10 consecutive rows
                    # Check average sentence length
                    segment_texts = call_df.loc[segment_indices, text_col].fillna('')
                    text_lengths = [len(str(text).strip()) for text in segment_texts]
                    avg_length = sum(text_lengths) / len(text_lengths) if text_lengths else 0
                    
                    if avg_length >= 20:  # Average length >= 20 characters
                        print(f"🎵 Salescall: Detected system recording - Speaker: {speaker}, "
                              f"Rows: {len(segment_indices)}, Avg length: {avg_length:.1f}")
                        # Add this speaker to the recording speakers set
                        recording_speakers.add(speaker)
                        # Break out of the segment loop since we've confirmed this speaker
                        break
        
        # Mark ALL lines from confirmed recording speakers as system recordings
        for speaker in recording_speakers:
            is_recording[call_df[speaker_col] == speaker] = True
            speaker_line_count = (call_df[speaker_col] == speaker).sum()
            print(f"📢 Salescall: Marking all {speaker_line_count} lines from {speaker} as system recordings")
        
        total_recording_rows = is_recording.sum()
        print(f"✅ Salescall: Detection complete - {total_recording_rows} rows marked as system recordings")
        
        return is_recording
    
    def _find_consecutive_segments(self, indices):
        """
        Salescall Helper: Find consecutive segments in a list of indices.
        
        Args:
            indices: List of row indices
            
        Returns:
            List of lists, each containing consecutive indices
        """
        if not indices:
            return []
        
        segments = []
        current_segment = [indices[0]]
        
        for i in range(1, len(indices)):
            if indices[i] == indices[i-1] + 1:  # Consecutive
                current_segment.append(indices[i])
            else:
                segments.append(current_segment)
                current_segment = [indices[i]]
        
        segments.append(current_segment)  # Add last segment
        return segments



    def build_grouped_lines(self, call_df, call_type="Sales Call", include_system_audio=False):
        """
        Unified grouped lines builder with configurable call type support.
        
        - call_type="Sales Call": Detects and excludes system audio recordings for grouping
        - call_type="SQCCB": Processes all rows without system recording detection
        - Uses three-pass physical merge strategy for optimal context grouping
        - Pass A: Forward merge for non-punctuated sentences
        - Pass B: Backward merge for short utterances  
        - Pass C: Forward merge for contextual windows
        """
        speaker_col, text_col = 'Speaker Roles', 'Transcription'
        start_col, end_col = 'Segment Start Time', 'Segment End Time'
        
        if speaker_col not in call_df.columns or text_col not in call_df.columns:
            raise KeyError(f"Required columns '{speaker_col}' and '{text_col}' not found.")
        
        # Create working copy to avoid modifying original DataFrame
        updated_call_df = call_df.copy()
        
        # Conditional system recording detection based on call_type
        if call_type == "Sales Call":
            # Sales Call mode: Detect system recordings
            system_recording_flags = self.detect_system_audio_recordings(call_df)
            updated_call_df['System_Audio_Recording'] = system_recording_flags.map({True: 'YES', False: 'NO'})
            
            # Apply system audio inclusion switch
            if include_system_audio:
                print(f"📊 Sales Call: Including all {len(call_df)} rows (including {system_recording_flags.sum()} system recordings) in analysis")
                analysis_df = updated_call_df.copy()
            else:
                print(f"📊 Sales Call: Filtering out {system_recording_flags.sum()} system recording rows from analysis")
                analysis_df = updated_call_df[~system_recording_flags].copy()
                print(f"📊 Sales Call: {len(analysis_df)} rows remaining for coverage analysis")
        else:
            # SQCCB mode: Skip system recording detection, process all rows
            print(f"🔍 SQCCB mode: Skip system recording detection, processing all {len(call_df)} rows")
            system_recording_flags = pd.Series([False] * len(call_df), index=call_df.index)
            updated_call_df['System_Audio_Recording'] = 'NO'
            analysis_df = updated_call_df.copy()

        def ends_with_punct(s):
            return s and s[-1] in '。！？；，.!?;'
        
        initial_rows = []
        for idx, row in analysis_df.iterrows():  # Use filtered DataFrame
            text = str(row.get(text_col, '')).strip()
            if not text: continue
            initial_rows.append({
                'orig_indices': [idx], 'text': text, 'speaker': row.get(speaker_col),
                'start_time': row.get(start_col), 'end_time': row.get(end_col),
            })

        if not initial_rows:
            print(f"⚠️  {call_type}: No valid rows found for grouping")
            return [], {}, updated_call_df, system_recording_flags

        # --- Pass A: Forward merge (non-punctuated) ---
        pass_a_results = []
        if initial_rows:
            pass_a_results.append(initial_rows[0].copy())
            for i in range(1, len(initial_rows)):
                current_row, last_group = initial_rows[i], pass_a_results[-1]
                if (last_group['speaker'] == current_row['speaker'] and 
                    not ends_with_punct(last_group['text']) and
                    (len(last_group['text']) + len(current_row['text']) + 1) <= 150):
                    last_group['text'] += ' ' + current_row['text']
                    last_group['orig_indices'].extend(current_row['orig_indices'])
                    if pd.notna(current_row['end_time']): last_group['end_time'] = current_row['end_time']
                else:
                    pass_a_results.append(current_row.copy())

        # --- Pass B: Backward merge (short utterances) ---
        pass_b_results = []
        if pass_a_results:
            for row in reversed(pass_a_results):
                if (pass_b_results and
                    pass_b_results[0]['speaker'] == row['speaker'] and
                    len(pass_b_results[0]['text']) < 20 and
                    (len(row['text']) + len(pass_b_results[0]['text']) + 1) <= 150):
                    first_group = pass_b_results[0]
                    first_group['text'] = row['text'] + ' ' + first_group['text']
                    first_group['orig_indices'] = row['orig_indices'] + first_group['orig_indices']
                    if pd.notna(row['start_time']): first_group['start_time'] = row['start_time']
                else:
                    pass_b_results.insert(0, row.copy())

        # --- Pass C: Forward contextual window merge ---
        final_results = []
        if pass_b_results:
            context_window_limit = 150  # Keep consistent with other limits
            final_results.append(pass_b_results[0].copy())
            for i in range(1, len(pass_b_results)):
                current_row, last_group = pass_b_results[i], final_results[-1]
                if (last_group['speaker'] == current_row['speaker'] and
                    (len(last_group['text']) + len(current_row['text']) + 1) <= context_window_limit):
                    last_group['text'] += ' ' + current_row['text']
                    last_group['orig_indices'].extend(current_row['orig_indices'])
                    if pd.notna(current_row['end_time']): last_group['end_time'] = current_row['end_time']
                else:
                    final_results.append(current_row.copy())

        # --- Final formatting ---
        grouped_lines, original_to_group = [], {}
        for group_id_counter, group in enumerate(final_results, 1):
            start_idx, end_idx = min(group['orig_indices']), max(group['orig_indices'])
            grouped_lines.append({
                'group_id': group_id_counter, 'text': group['text'], 'speaker': group['speaker'],
                'start_idx': start_idx, 'end_idx': end_idx,
                'start_time': group['start_time'], 'end_time': group['end_time'],
                'orig_indices': group['orig_indices']  # ✅ 添加orig_indices字段
            })
            for orig_idx in group['orig_indices']:
                original_to_group[orig_idx] = group_id_counter
            
        print(f"🔀 {call_type} grouping completed: {len(grouped_lines)} final groups created")
        
        # Return enhanced results: grouped_lines, original_to_group, updated_call_df, system_recording_flags
        return grouped_lines, original_to_group, updated_call_df, system_recording_flags
    
    def get_business_overlapping_keywords(self, text1, text2):
        """
        Get overlapping keywords focusing on business terms only.
        Filters out common words and focuses on important business terms.
        """
        tokens1 = set(self.tokenize_text(text1))
        tokens2 = set(self.tokenize_text(text2))
        
        # Get all overlapping tokens
        all_overlap = tokens1 & tokens2
        
        # Prioritize important business terms
        important_overlap = all_overlap & self.important_keywords
        
        # Return important terms first, then other business terms
        result = list(important_overlap) + [token for token in all_overlap if token not in important_overlap]
        return result

    def calculate_keyword_coverage(self, script_text, group_text):
        """
        Calculate keyword coverage in [0,1]: coverage = |H ∩ K| / |K|
        K: key business terms from script (tokens ∩ important_keywords)
        H: tokens from group text
        """
        script_tokens = set(self.tokenize_text(self.preprocess_script_text(script_text)))  # script text
        group_tokens = set(self.tokenize_text(self.preprocess_call_text(group_text)))      # call text
        K = script_tokens & self.important_keywords
        if not K:
            return 0.0
        H = group_tokens & K
        return float(len(H) / len(K))

    # ------------------------------------------------------
    # Enhanced Pattern Recognition for Specific Discussion Points
    # ------------------------------------------------------
    def setup_enhanced_scoring_config(self):
        """
        Load enhanced scoring configuration from USER_ENHANCED_DISCUSSION_POINTS defined at top of file.
        
        CONFIGURATION LOCATION:
        ======================
        The discussion points configuration is now at the TOP of this file.
        Search for: "USER_ENHANCED_DISCUSSION_POINTS" 
        
        WHEN IS THIS TRIGGERED:
        =======================
        This enhancement is automatically applied during similarity calculation
        in _analyze_single_point_coverage when comparing script variations against call text.
        The system will check if the discussion point name matches any key in USER_ENHANCED_DISCUSSION_POINTS.
        
        HOW IT WORKS:
        =============
        1. System calculates base similarity score (0-1) using TF-IDF, synonyms, etc.
        2. If discussion point is in USER_ENHANCED_DISCUSSION_POINTS, checks for patterns
        3. If date patterns found AND date_boost > 0: adds date_boost to score
        4. If numeric patterns found AND numeric_boost > 0: adds numeric_boost to score  
        5. Final score is capped at 1.0
        """
        return USER_ENHANCED_DISCUSSION_POINTS
    
    def detect_date_patterns(self, text):
        """
        Detect various Chinese date formats in text (optimized for performance).
        Supports mixed formats like: 2025年十月九日, 二零二五年10月9日, etc.
        """
        # Quick pre-check to avoid regex if no date indicators
        if not any(indicator in text for indicator in ['年', '月', '日', '/', '-']):
            return False, []
        
        # Comprehensive date patterns (mixed Chinese/Arabic numerals)
        patterns = [
            r'[二零一九八七六五四三]{2,4}年[一二三四五六七八九十\d]{1,3}月[一二三四五六七八九十\d]{1,3}日',  # Mixed format: 二零二五年10月9日
            r'\d{4}年[一二三四五六七八九十]{1,3}月[一二三四五六七八九十]{1,3}日',  # Mixed format: 2025年十月九日
            r'\d{4}年\d{1,2}月\d{1,2}日',  # Standard: 2025年10月9日
            r'[一二三四五六七八九十]{1,3}月[一二三四五六七八九十\d]{1,3}日',  # Month-Day format: 十月九日, 十月9日
            r'\d{1,2}月\d{1,2}日',  # Simple: 10月9日
            r'\d{4}[-/]\d{1,2}[-/]\d{1,2}',  # Western: 2025-10-09
        ]
        
        found_dates = []
        for pattern in patterns:
            matches = re.findall(pattern, text)
            found_dates.extend(matches)
            if found_dates:  # Early exit optimization
                break
                
        return len(found_dates) > 0, found_dates[:3]  # Limit to first 3 for performance
    
    def detect_numeric_patterns(self, text):
        """
        Detect numeric patterns excluding dates (optimized for performance).
        Focuses on prices, amounts, percentages, and financial numbers.
        """
        # Quick pre-check for numeric content
        if not re.search(r'[\d￥$¥€£%万千百亿]', text):
            return False, []
        
        # High-priority patterns (most common in financial calls)
        priority_patterns = [
            r'[港人民币]{1,3}\s*\d+(?:\.\d+)?[万千百亿]?',  # Currency: 港币50万
            r'[￥$¥€£]\s*[\d,]+(?:\.\d+)?',  # $100,000
            r'\d+(?:\.\d+)?[万千百亿]',  # Amount: 50万
            r'\d+(?:\.\d+)?%',  # 15.5%
        ]
        
        found_numbers = []
        for pattern in priority_patterns:
            matches = re.findall(pattern, text)
            if matches:
                found_numbers.extend(matches)
                break  # Early exit for performance
        
        # Fallback: large numbers (but exclude date-like patterns)
        if not found_numbers:
            large_nums = re.findall(r'\d{4,}', text)
            for num in large_nums:
                # Skip if looks like year or date component
                if not (1900 <= int(num) <= 2100 or len(num) == 4):
                    found_numbers.append(num)
                    break  # Only take first one for performance
        
        return len(found_numbers) > 0, found_numbers[:2]  # Limit for performance
    
    def apply_pattern_enhancement(self, base_score, text, discussion_point, dynamic_numeric_boost=0):
        """
        Apply pattern-based score enhancement for specific discussion points.
        This is called automatically during similarity calculation.
        
        Args:
            base_score: Original weighted similarity score (0-1)
            text: Call text to analyze for patterns  
            discussion_point: Name of discussion point being analyzed
            dynamic_numeric_boost: Dynamic numeric boost value (optional)
            
        Returns:
            Enhanced score (capped at 1.0)
        """
        # Check if this discussion point has enhancement configured
        enhancement_config = self.setup_enhanced_scoring_config()
        enhanced_score = base_score
        
        # Priority 1: Use static configuration if available
        if discussion_point in enhancement_config:
            config = enhancement_config[discussion_point]
            
            # Apply date enhancement
            if config.get('date_boost', 0) > 0:
                has_dates, _ = self.detect_date_patterns(text)
                if has_dates:
                    enhanced_score += config['date_boost']
            
            # Apply numeric enhancement  
            if config.get('numeric_boost', 0) > 0:
                has_numbers, _ = self.detect_numeric_patterns(text)
                if has_numbers:
                    enhanced_score += config['numeric_boost']
        
        # Priority 2: Use dynamic numeric boost if no static config and dynamic boost is available
        elif dynamic_numeric_boost > 0:
            has_numbers, _ = self.detect_numeric_patterns(text)
            if has_numbers:
                enhanced_score += dynamic_numeric_boost
        
        return min(1.0, enhanced_score)  # Cap at 1.0

    # ------------------------------------------------------
    # Advanced Semantic Similarity Calculation
    # ------------------------------------------------------
    def calculate_semantic_similarity(self, text1, text2):
        """
        Calculate semantic similarity using updated metric system:
        1. Business-weighted TF-IDF + Cosine Similarity with char n-grams (2,4) (55%)
        2. Expanded token overlap with synonyms (15%)
        3. Token-level ROUGE-L (LCS-based order sensitivity) (20%)
        Returns a dictionary of all metrics. Business/exact-term boosts applied separately.
        """
        # Clean and preprocess both texts
        clean1 = self.preprocess_script_text(text1)  # text1 is script text
        clean2 = self.preprocess_call_text(text2)    # text2 is call text
        
        metrics = {}
        
        # 1. Business-weighted TF-IDF + Cosine Similarity (with char n-grams)
        tfidf_cosine_score = self.calculate_business_weighted_tfidf_cosine(clean1, clean2)
        metrics['tfidf_cosine'] = tfidf_cosine_score
        
        # 2. Expanded token overlap (with synonyms)
        expanded1 = self.expand_keywords(clean1)
        expanded2 = self.expand_keywords(clean2)
        expanded_overlap = len(expanded1 & expanded2)
        expanded_similarity = expanded_overlap / max(len(expanded1), len(expanded2)) if max(len(expanded1), len(expanded2)) > 0 else 0
        metrics['expanded_overlap'] = expanded_similarity
        
        # 3. Token-level ROUGE-L (LCS-based sequence similarity)
        rouge_l = self.calculate_token_rouge_l(clean1, clean2)
        metrics['rouge_l'] = rouge_l

        # 4. Keyword coverage (bounded 0..1)
        keyword_coverage = self.calculate_keyword_coverage(text1, text2)
        metrics['keyword_coverage'] = keyword_coverage
        
        # Weighted combination (Optimization 1)
        weighted_score = (
            tfidf_cosine_score * 0.55 +      # TF-IDF + Cosine with char n-grams
            expanded_similarity * 0.15 +     # Synonym expansion
            rouge_l * 0.20 +                 # Char-level ROUGE-L
            keyword_coverage * 0.10          # Bounded keyword coverage
        )
        metrics['weighted_score'] = weighted_score
        
        return metrics
    
    def _get_or_create_tfidf_vectorizer(self, corpus_sample):
        """
        Get cached TF-IDF vectorizer or create a new one if needed.
        Reuses vectorizer across multiple text comparisons for efficiency.
        """
        if self._tfidf_vectorizer is None:
            # Create TF-IDF vectorizer with char n-grams (2,4)
            self._tfidf_vectorizer = TfidfVectorizer(
                lowercase=False,  # Preserve Chinese characters
                analyzer='char',
                use_idf=True,
                smooth_idf=True,
                sublinear_tf=True,  # Use log scaling for term frequency
                ngram_range=(2, 4),
                max_features=50000  # Limit features to control memory usage
            )
            # Fit on a sample corpus to establish vocabulary
            self._tfidf_vectorizer.fit(corpus_sample)
            
        return self._tfidf_vectorizer

    def calculate_business_weighted_tfidf_cosine(self, text1, text2):
        """
        Calculate business-weighted TF-IDF + Cosine Similarity (optimized).
        Uses cached vectorizer and keeps matrices sparse to save memory.
        
        Returns:
        - Float: Similarity score between 0 and 1
        """
        # Check similarity cache first
        cache_key = (text1, text2)
        if cache_key in self._similarity_cache:
            return self._similarity_cache[cache_key]
        
        # Tokenize texts (already filtered for business relevance)
        tokens1 = self.tokenize_text(text1)
        tokens2 = self.tokenize_text(text2)
        
        # Handle edge cases
        if not tokens1 or not tokens2:
            return 0.0
        
        # Create document corpus
        doc1 = ' '.join(tokens1)
        doc2 = ' '.join(tokens2)
        corpus = [doc1, doc2]
        
        try:
            # Get cached vectorizer or create new one
            vectorizer = self._get_or_create_tfidf_vectorizer(corpus)
            
            # Transform the corpus (reuse fitted vectorizer)
            tfidf_matrix = vectorizer.transform(corpus)
            
            # Apply tiered dynamic weights to TF-IDF features (keep sparse)
            feature_names = vectorizer.get_feature_names_out()
            weighted_matrix = self.apply_tiered_weights_to_tfidf_sparse(tfidf_matrix, feature_names)
            
            # Calculate cosine similarity
            cosine_sim = cosine_similarity(weighted_matrix[0:1], weighted_matrix[1:2])[0][0]
            
            result = float(cosine_sim)
            
            # Cache result (limit cache size)
            if len(self._similarity_cache) < 5000:
                self._similarity_cache[cache_key] = result
                
            return result
            
        except Exception as e:
            # Fallback to simple token overlap if TF-IDF fails
            tokens1_set = set(tokens1)
            tokens2_set = set(tokens2)
            overlap = len(tokens1_set & tokens2_set)
            union = len(tokens1_set | tokens2_set)
            return overlap / union if union > 0 else 0.0
    
    def apply_tiered_weights_to_tfidf_sparse(self, tfidf_matrix, feature_names):
        """
        Apply Tiered Dynamic Weighting to TF-IDF feature columns (sparse version).
        This version keeps the matrix sparse to save memory.
        """
        import math
        from scipy import sparse
        
        # Build sets/dicts for fast lookup
        stop2to4 = {w for w in self.stopwords if isinstance(w, str) and 2 <= len(w) <= 4}
        core2to4 = {w: self.current_weights[w] for w in self.current_weights if isinstance(w, str) and 2 <= len(w) <= 4}
        freq = self.call_char_ngram_freq or {}
        
        # Create weight vector
        weights = []
        for token in feature_names:
            if token in stop2to4:
                w = 0.0
            elif token in core2to4:
                f = freq.get(token, 0)
                discount = 1.0 / (1.0 + math.log(1.0 + f)) if f > 0 else 1.0
                w = core2to4[token] * discount
            else:
                f = freq.get(token, 0)
                w = 0.4 if f > self.high_freq_threshold else 1.0
            weights.append(w)
        
        # Apply weights by element-wise multiplication (keeps sparse)
        weight_matrix = sparse.diags(weights, format='csr')
        weighted_matrix = tfidf_matrix @ weight_matrix
        
        return weighted_matrix
    
    def apply_tiered_weights_to_tfidf(self, tfidf_matrix, feature_names):
        """
        Apply Tiered Dynamic Weighting to TF-IDF feature columns (legacy dense version).
        Note: This function is kept for backward compatibility but not used in optimized path.
        """
        dense = tfidf_matrix.toarray()
        # Build sets/dicts for fast lookup
        stop2to4 = {w for w in self.stopwords if isinstance(w, str) and 2 <= len(w) <= 4}
        core2to4 = {w: self.current_weights[w] for w in self.current_weights if isinstance(w, str) and 2 <= len(w) <= 4}
        freq = self.call_char_ngram_freq or {}

        import math
        for idx, token in enumerate(feature_names):
            if token in stop2to4:
                w = 0.0
            elif token in core2to4:
                f = freq.get(token, 0)
                discount = 1.0 / (1.0 + math.log(1.0 + f)) if f > 0 else 1.0
                w = core2to4[token] * discount
            else:
                f = freq.get(token, 0)
                w = 0.4 if f > self.high_freq_threshold else 1.0
            if w == 0.0:
                dense[:, idx] = 0.0
            elif w != 1.0:
                dense[:, idx] *= w
        return dense

    def precompute_call_char_ngram_freq(self, call_df):
        """
        Precompute char 2-4 gram frequencies over the entire call text once per run.
        """
        text_col = 'Transcription'
        if text_col not in call_df.columns:
            self.call_char_ngram_freq = {}
            return
        from collections import Counter
        counter = Counter()
        # Iterate rows and count ngrams on cleaned text
        for _, row in call_df.iterrows():
            raw = row[text_col]
            if raw is None or pd.isna(raw):
                continue
            s = self.preprocess_call_text(str(raw))  # call text
            if not s:
                continue
            L = len(s)
            for n in (2, 3, 4):
                if L < n:
                    continue
                for i in range(L - n + 1):
                    ngram = s[i:i+n]
                    counter[ngram] += 1
        self.call_char_ngram_freq = dict(counter)

    def calculate_token_rouge_l(self, text1, text2):
        """
        Compute character-level ROUGE-L F-score using LCS over character sequences.
        For Chinese/Cantonese, char-level LCS is robust to tokenization errors.
        Returns value in [0,1].
        """
        seq1 = list(text1)
        seq2 = list(text2)
        if not seq1 or not seq2:
            return 0.0
        n, m = len(seq1), len(seq2)
        dp = [[0]*(m+1) for _ in range(n+1)]
        for i in range(1, n+1):
            c1 = seq1[i-1]
            for j in range(1, m+1):
                if c1 == seq2[j-1]:
                    dp[i][j] = dp[i-1][j-1] + 1
                else:
                    dp[i][j] = dp[i-1][j] if dp[i-1][j] >= dp[i][j-1] else dp[i][j-1]
        lcs_len = dp[n][m]
        if lcs_len == 0:
            return 0.0
        precision = lcs_len / m
        recall = lcs_len / n
        if precision + recall == 0:
            return 0.0
        f_score = (2 * precision * recall) / (precision + recall)
        return float(f_score)
    
    # Note: Removed calculate_business_term_density and calculate_keyword_importance
    # These functions have been replaced by the business-weighted TF-IDF + Cosine Similarity approach
    # which inherently handles term importance and business relevance more effectively.
    # 
    # Note: Dynamic term importance analysis has been moved to dynamic_script_analysis.py
    # This keeps the main analysis tool focused on coverage checking.

    # ------------------------------------------------------
    # Script Variation Parsing
    # ------------------------------------------------------
    def parse_script_variations(self, script_text):
        """
        Parse multiple script variations from a single script text.
        Enhanced with preprocessing and comprehensive short fragment merging.
        Handles different customer categories (version A, B, etc.) and multiple sentences.
        
        Args:
        - script_text: Raw script text containing multiple variations/sentences
        
        Returns:
        - List of individual script variations
        """
        if not script_text or pd.isna(script_text):
            return []
        
        script_text = str(script_text).strip()
        if not script_text:
            return []
        
        # Step 1: 先预处理去英文
        preprocessed_text = self.preprocess_text(script_text, mode='comparison')
        if not preprocessed_text:
            return []
        
        # Step 2: 第一次分割
        raw_splits = re.split(r'[.;\n]|版本[AB][:：]|Version [AB][:：]|情況[一二三四五六七八九十][:：]', preprocessed_text)
        
        # Step 3: 第一次短片段合并
        cleaned_splits = [split.strip() for split in raw_splits if split.strip()]
        merged_splits = self._merge_short_fragments(cleaned_splits)
        
        # Step 4: 第二次分割（处理过长片段）
        variations = []
        for split in merged_splits:
            if len(split) > 50:
                # 按逗号分割
                comma_splits = [s.strip() for s in split.split(',') if s.strip()]
                # 第二次短片段合并
                comma_merged = self._merge_short_fragments(comma_splits)
                variations.extend(comma_merged)
            else:
                variations.append(split)
        
        # Step 5: 去重
        unique_variations = []
        seen = set()
        for var in variations:
            if var not in seen:
                unique_variations.append(var)
                seen.add(var)
        
        # If no variations found, return the preprocessed text as single variation
        if not unique_variations:
            unique_variations = [preprocessed_text]
        
        return unique_variations
    
    def _merge_short_fragments(self, fragments):
        """
        Helper function: 合并≤5字的短片段
        短片段向后合并，最后一个向前合并
        
        Args:
            fragments: List of text fragments
            
        Returns:
            List of merged fragments with no fragment ≤5 characters
        """
        if not fragments:
            return []
        
        merged = []
        i = 0
        while i < len(fragments):
            current = fragments[i]
            
            if len(current) <= 5:
                # 尝试向后合并
                if i + 1 < len(fragments):
                    fragments[i + 1] = current + ' ' + fragments[i + 1]
                # 如果是最后一个，向前合并
                elif merged:
                    merged[-1] = merged[-1] + ' ' + current
                else:
                    merged.append(current)  # 只有这一个片段
            else:
                merged.append(current)
            i += 1
        
        return merged

    # ------------------------------------------------------
    # Coverage Checking Logic
    # ------------------------------------------------------
    def _initialize_tfidf_with_corpus(self, call_df, required_points_df):
        """
        Initialize TF-IDF vectorizer with all available text to establish vocabulary.
        This is called once at the beginning of coverage checking.
        """
        if self._tfidf_vectorizer is not None:
            return  # Already initialized
            
        # Collect all text for vocabulary building
        corpus_texts = []
        
        # Add call texts
        text_col = 'Transcription'
        if text_col in call_df.columns:
            for _, row in call_df.iterrows():
                text = row[text_col]
                if not pd.isna(text) and str(text).strip():
                    tokens = self.tokenize_text(self.preprocess_call_text(str(text)))  # call text
                    if tokens:
                        corpus_texts.append(' '.join(tokens))
        
        # Add script texts
        for _, row in required_points_df.iterrows():
            script = row['Standard_Script']
            if not pd.isna(script):
                variations = self.parse_script_variations(str(script))
                for variation in variations:
                    tokens = self.tokenize_text(self.preprocess_script_text(variation))  # script variation
                    if tokens:
                        corpus_texts.append(' '.join(tokens))
        
        # Initialize vectorizer with all corpus
        if corpus_texts:
            self._get_or_create_tfidf_vectorizer(corpus_texts)

    def check_coverage(self, call_df, required_points_df, grouped_lines, threshold=0.3):
        """
        Check if call text covers all required discussion points using provided grouped_lines.
        - Takes pre-computed grouped_lines to avoid re-calculation
        - For each required point, collects ALL script rows for that point and parses multiple variations.
        - Compares ALL variations for the point against each speaker group.
        - Takes the highest score among all variations as the final score for each point.
        - Mark as 'Covered' if highest score >= threshold.
        - Returns DataFrame with detailed results including matched groups and best variation.
        """
        # Initialize TF-IDF vectorizer once with all corpus text
        self._initialize_tfidf_with_corpus(call_df, required_points_df)
        
        results = []
        
        # Group scripts by Required_Discussion_Point
        point_to_scripts = {}
        for _, row in required_points_df.iterrows():
            p = row['Required_Discussion_Point']
            s = row['Standard_Script']
            if pd.isna(p) or pd.isna(s):
                continue
            point_to_scripts.setdefault(p, []).append(str(s))

        for point, scripts in point_to_scripts.items():
            result = self._analyze_single_point_coverage(point, scripts, grouped_lines, threshold)
            results.append(result)
        
        return pd.DataFrame(results)
    
    def _analyze_single_point_coverage(self, point, scripts, grouped_lines, threshold):
        """
        Analyze coverage for a single discussion point using grouped_lines.
        NOW INCLUDES ENHANCED PATTERN RECOGNITION for configured discussion points.
        Uses point-specific weights for more precise analysis.
        IMPLEMENTS DUAL VERIFICATION: Holistic vs Granular matching strategy.
        """
        script_variations = [var for s in scripts for var in self.parse_script_variations(s)]
        
        # Create holistic script text by joining all scripts with spaces
        holistic_script_text = ' '.join(scripts) if scripts else ''
        
        best_score = 0
        best_group_info, best_variation, best_metrics = {}, "", {}
        best_original_score = 0  # Track original score before enhancement
        best_match_type = ""  # Track whether final match was Holistic or Granular
        best_holistic_score = 0  # Track holistic score for the best match
        best_granular_score = 0  # Track granular score for the best match
        
        # Dynamic numeric boost detection
        dynamic_numeric_boost = 0
        
        # Dynamic rule 1: Check point name for price/floating rate keywords
        point_lower = point.lower()
        if 'price' in point_lower or 'floating rate' in point_lower:
            dynamic_numeric_boost = 0.15
        
        # Dynamic rule 2: Check script content for price/interest/% keywords
        if dynamic_numeric_boost == 0 and scripts:
            combined_scripts_lower = ' '.join(scripts).lower()
            if any(keyword in combined_scripts_lower for keyword in ['price', 'interest', '%']):
                dynamic_numeric_boost = 0.15
        
        # Get point-specific weights for this discussion point
        try:
            point_weights = dictionaries.get_point_specific_weights(
                self.current_language, self.current_product, point
            )
            # Temporarily switch to point-specific weights
            original_weights = self.current_weights
            self.current_weights = point_weights
            point_specific_mode = True
        except Exception as e:
            # Fallback to existing weights if point-specific weights fail
            point_specific_mode = False
        
        for group_info in grouped_lines:
            group_text = group_info['text']
            if not group_text: continue
            
            # Step 1: Holistic matching - compare with complete script text
            holistic_metrics = self.calculate_semantic_similarity(holistic_script_text, group_text)
            holistic_weighted_score = holistic_metrics['weighted_score']
            
            # Step 2: Granular matching - find best among script variations
            best_granular_score_for_group = 0
            best_granular_metrics_for_group = {}
            best_variation_for_group = ""
            
            for variation in script_variations:
                # Calculate base similarity metrics (now uses point-specific weights)
                metrics = self.calculate_semantic_similarity(variation, group_text)
                granular_score = metrics['weighted_score']
                
                if granular_score > best_granular_score_for_group:
                    best_granular_score_for_group = granular_score
                    best_granular_metrics_for_group = metrics
                    best_variation_for_group = variation
            
            # Step 3: Determine final score and match type for this group
            if holistic_weighted_score > best_granular_score_for_group:
                final_score_for_group = holistic_weighted_score
                final_metrics_for_group = holistic_metrics
                match_type_for_group = "Holistic"
                final_variation_for_group = holistic_script_text
            else:
                final_score_for_group = best_granular_score_for_group
                final_metrics_for_group = best_granular_metrics_for_group
                match_type_for_group = "Granular"
                final_variation_for_group = best_variation_for_group
            
            # Apply pattern enhancement to the final score
            enhanced_score = self.apply_pattern_enhancement(final_score_for_group, group_text, point, dynamic_numeric_boost)
            
            # Step 4: Update global best if this group has higher score
            if enhanced_score > best_score:
                best_score = enhanced_score
                best_original_score = final_score_for_group
                best_group_info, best_metrics, best_variation = group_info, final_metrics_for_group, final_variation_for_group
                best_match_type = match_type_for_group
                best_holistic_score = holistic_weighted_score
                best_granular_score = best_granular_score_for_group
                # Update metrics with enhanced score
                best_metrics['weighted_score'] = enhanced_score
                if enhanced_score >= 0.9: break
            if best_score >= 0.9: break
        
        # Restore original weights if we switched to point-specific mode
        if point_specific_mode:
            self.current_weights = original_weights
        
        overlapping_keywords = self.get_business_overlapping_keywords(self.preprocess_script_text(best_variation or (scripts[0] if scripts else "")), best_group_info.get('text',''))  # script variation
        
        # Calculate enhancement boost for reporting
        enhancement_boost = best_score - best_original_score
        
        return {
            'Required_Discussion_Point': point, 'Covered': 'Covered' if best_score >= threshold else 'Not Covered',
            'Weighted_Score': round(best_score, 3), 'Comparison_Score': round(best_score, 3),
            'Original_Score': round(best_original_score, 3),  # New: show original score
            'Enhancement_Boost': round(enhancement_boost, 3),  # New: show boost amount
            'Match_Type': best_match_type,  # New: Holistic or Granular
            'Holistic_Score': round(best_holistic_score, 3),  # New: holistic matching score
            'Granular_Score': round(best_granular_score, 3),  # New: granular matching score
            'TF_IDF_Cosine': round(best_metrics.get('tfidf_cosine', 0), 3),
            'Expanded_Overlap': round(best_metrics.get('expanded_overlap', 0), 3),
            'ROUGE_L': round(best_metrics.get('rouge_l', 0), 3),
            'Keyword_Coverage': round(best_metrics.get('keyword_coverage', 0), 3),
            'Overlapping_Keywords': ', '.join(overlapping_keywords), 'Matched_Group': best_group_info.get('text', ''),
            'Group_ID': best_group_info.get('group_id', -1), 'Speaker': best_group_info.get('speaker', ''),
            'Best_Matching_Variation': best_variation or (scripts[0] if scripts else ''),
            'All_Variations_Count': len(script_variations)
        }


    def create_grouped_call_dataframe(self, grouped_lines):
        """Creates a DataFrame from the final grouped lines for reference."""
        return pd.DataFrame(grouped_lines)
    
    def create_call_text_analysis_view(self, grouped_lines, script_df, updated_call_df, threshold=0.3):
        """
        Creates a call text analysis view with full functionality (_Hit, _Score, _Keywords).
        This version includes all groups (including system recordings) and adds system recording identification.
        """
        analysis_rows = []
        point_to_variations = {}
        for _, row in script_df.iterrows():
            p, s = row['Required_Discussion_Point'], row['Standard_Script']
            if pd.notna(p) and pd.notna(s):
                point_to_variations.setdefault(p, []).extend(self.parse_script_variations(str(s)))
        
        all_points = list(point_to_variations.keys())

        for group in grouped_lines:
            analysis_row = {'Group_ID': group['group_id'], 'Speaker': group['speaker'], 'Call_Text': group['text']}
            
            # Check if this group is a system recording by examining the first original index
            first_orig_index = group['orig_indices'][0] if group['orig_indices'] else None
            is_system_recording = False
            if first_orig_index is not None and first_orig_index < len(updated_call_df):
                is_system_recording = updated_call_df.iloc[first_orig_index]['System_Audio_Recording']
            
            # Add system recording identification column
            analysis_row['System_Audio_Recording'] = 'YES' if is_system_recording else 'NO'
            
            # Process discussion points
            for point in all_points:
                point_short = f"{point[:30]}{'...' if len(point) > 30 else ''}"
                
                if is_system_recording:
                    # For system recordings, set default values without computation
                    analysis_row[f"{point_short}_Hit"] = 'NO'
                    analysis_row[f"{point_short}_Score"] = 0.0
                    analysis_row[f"{point_short}_Keywords"] = ''
                else:
                    # For non-system recordings, perform full similarity calculation
                    script_variations = point_to_variations.get(point, [])
                    best_score_for_point = 0
                    best_keywords_for_point = []
                    
                    for variation in script_variations:
                        similarity_metrics = self.calculate_semantic_similarity(variation, group['text'])
                        score = similarity_metrics['weighted_score']
                        if score > best_score_for_point:
                            best_score_for_point = score
                        best_keywords_for_point = self.get_business_overlapping_keywords(
                            self.preprocess_script_text(variation), group['text']  # script variation
                        )
                    
                    point_hit = best_score_for_point >= threshold
                    analysis_row[f"{point_short}_Hit"] = 'YES' if point_hit else 'NO'
                    analysis_row[f"{point_short}_Score"] = round(best_score_for_point, 3)
                    analysis_row[f"{point_short}_Keywords"] = ', '.join(best_keywords_for_point)
            
            analysis_rows.append(analysis_row)
        
        return pd.DataFrame(analysis_rows)

    def create_call_text_analysis_view_with_separation(self, grouped_lines, system_audio_df, script_df, threshold=0.3):
        """
        STEP 4: Late Merge & Reporting - Create call text analysis with proper separation handling.
        
        This function implements the late merge strategy:
        1. Generate analysis results for human dialogue groups (grouped_lines) - includes Hit, Score, Keywords
        2. Format system recording rows with minimal structure (_Hit='NO' only, no Score/Keywords)
        3. Combine both into a single comprehensive report
        
        Args:
            grouped_lines: List of grouped human dialogue segments
            system_audio_df: DataFrame containing only system recording rows
            script_df: DataFrame with script data
            threshold: Threshold for determining hits
            
        Returns:
            pd.DataFrame: Combined analysis view with human dialogue analysis + system recording rows
        """
        analysis_rows = []
        
        # Group script variations by discussion point
        point_to_variations = {}
        for _, row in script_df.iterrows():
            point = row['Required_Discussion_Point']
            script = row['Standard_Script']
            if pd.notna(point) and pd.notna(script):
                point_to_variations.setdefault(point, []).extend(self.parse_script_variations(str(script)))
        
        all_points = list(point_to_variations.keys())
        
        # Part 1: Process human dialogue groups (with actual analysis)
        for group in grouped_lines:
            analysis_row = {
                'Group_ID': group['group_id'], 
                'Speaker': group['speaker'], 
                'Call_Text': group['text'],
                'System_Audio_Recording': 'NO'  # All human dialogue
            }
            
            # Analyze each discussion point for this group
            for point in all_points:
                point_short = f"{point[:30]}{'...' if len(point) > 30 else ''}"
                script_variations = point_to_variations.get(point, [])
                
                best_score_for_point = 0
                best_keywords_for_point = []
                
                # Analyze against all variations for this point
                for variation in script_variations:
                    metrics = self.calculate_semantic_similarity(variation, group['text'])
                    original_score = metrics['weighted_score']
                    enhanced_score = self.apply_pattern_enhancement(original_score, group['text'], point)
                    
                    if enhanced_score > best_score_for_point:
                        best_score_for_point = enhanced_score
                        best_keywords_for_point = self.get_business_overlapping_keywords(
                            self.preprocess_script_text(variation), group['text']
                        )
                
                # Record results for this point
                point_hit = best_score_for_point >= threshold
                analysis_row[f"{point_short}_Hit"] = 'YES' if point_hit else 'NO'
                analysis_row[f"{point_short}_Score"] = round(best_score_for_point, 3)
                analysis_row[f"{point_short}_Keywords"] = ', '.join(best_keywords_for_point)
            
            analysis_rows.append(analysis_row)
        
        # Part 2: Process system recording rows (only Hit column needed)
        if not system_audio_df.empty:
            for idx, row in system_audio_df.iterrows():
                system_row = {
                    'Group_ID': f"SYS_{idx}",  # Unique ID for system recordings
                    'Speaker': row.get('Speaker Roles', 'System'),
                    'Call_Text': str(row.get('Transcription', '')),
                    'System_Audio_Recording': 'YES'
                }
                
                # For system recordings, only Hit column is needed
                for point in all_points:
                    point_short = f"{point[:30]}{'...' if len(point) > 30 else ''}"
                    system_row[f"{point_short}_Hit"] = 'NO'
                
                analysis_rows.append(system_row)
        
        # Part 3: Create final DataFrame
        final_df = pd.DataFrame(analysis_rows)
        
        return final_df

    def create_sentence_level_output(self, call_df, original_to_group):
        """Creates sentence-level output mapping original rows to their final group ID."""
        speaker_col, text_col = 'Speaker Roles', 'Transcription'
        start_col, end_col = 'Segment Start Time', 'Segment End Time'
        
        sentence_level_data = []
        for idx, row in call_df.iterrows():
            text_val = row.get(text_col)
            if pd.isna(text_val) or not str(text_val).strip(): continue
            sentence_level_data.append({
                'Original_Row': idx, 'Group_ID': original_to_group.get(idx),
                'Speaker': row.get(speaker_col, ''), 'Start_Time': row.get(start_col),
                'End_Time': row.get(end_col), 'Original_Text': str(text_val).strip()
            })
        return pd.DataFrame(sentence_level_data)
    
    def parse_text_into_sentences(self, text):
        """
        Parse text into individual sentences using punctuation and natural breaks.
        """
        # Split by sentence-ending punctuation
        sentence_pattern = r'[。！？.!?]+'
        sentences = re.split(sentence_pattern, text)
        
        # Clean up sentences and filter out empty ones
        cleaned_sentences = []
        for sentence in sentences:
            sentence = sentence.strip()
            if sentence and len(sentence) > 2:  # Filter out very short fragments
                cleaned_sentences.append(sentence)
        
        # If no sentences found, try splitting by commas for very long text
        if not cleaned_sentences and len(text) > 100:
            comma_splits = text.split('，')
            for split in comma_splits:
                split = split.strip()
                if split and len(split) > 10:
                    cleaned_sentences.append(split)
        
        # If still no sentences, return the original text as one sentence
        if not cleaned_sentences:
            cleaned_sentences = [text]
        
        return cleaned_sentences

# ==========================================================
# Helper Functions for Executive Summary
# ==========================================================
def parse_time_to_seconds(time_str):
    """
    Parse time string in format "00:18:00" to total seconds.
    Extracts the last 6 characters (MM:SS) and converts to seconds.
    """
    if pd.isna(time_str) or not time_str:
        return 0
    
    try:
        time_str = str(time_str).strip()
        # Extract last 6 characters for MM:SS format
        if len(time_str) >= 6:
            time_part = time_str[-6:]  # Get "18:00" from "00:18:00"
        else:
            time_part = time_str
        
        # Split and convert to seconds
        if ':' in time_part:
            parts = time_part.split(':')
            if len(parts) == 2:
                minutes = int(parts[0])
                seconds = int(parts[1])
                return minutes * 60 + seconds
            elif len(parts) == 3:  # Handle full HH:MM:SS
                hours = int(parts[0])
                minutes = int(parts[1])
                seconds = int(parts[2])
                return hours * 3600 + minutes * 60 + seconds
        
        # If no colon, try to parse as plain number
        return float(time_str)
    except (ValueError, TypeError):
        return 0


def apply_composite_point_key_if_configured(script_df, script_sheet_name):
    """
    For configured sheets only, combine:
      Required_Discussion_Point + Points to Note
    as the new Required_Discussion_Point.
    """
    if not should_use_composite_point_key(script_sheet_name):
        return script_df

    required_col = 'Required_Discussion_Point'
    if required_col not in script_df.columns:
        print(f"⚠️  {script_sheet_name}: Missing '{required_col}', skip composite point key")
        return script_df

    note_col = None
    for col in script_df.columns:
        if str(col).strip().lower() == str(COMPOSITE_POINT_NOTE_COLUMN).strip().lower():
            note_col = col
            break

    if note_col is None:
        print(f"⚠️  {script_sheet_name}: Missing '{COMPOSITE_POINT_NOTE_COLUMN}', skip composite point key")
        return script_df

    updated_df = script_df.copy()

    def _combine(row):
        base = '' if pd.isna(row.get(required_col)) else str(row.get(required_col)).strip()
        note = '' if pd.isna(row.get(note_col)) else str(row.get(note_col)).strip()
        if note:
            return f"{base}{COMPOSITE_POINT_KEY_SEPARATOR}{note}" if base else note
        return base

    updated_df[required_col] = updated_df.apply(_combine, axis=1)
    print(f"🧩 {script_sheet_name}: Composite point key enabled ({required_col} + {note_col})")
    return updated_df

def generate_executive_summary_report(results_df, updated_call_df, script_df, grouped_lines, output_file=None):
    """
    Generate a consolidated Executive Summary Excel report with three sections:
    A: Overall Performance KPIs
    B: Key Insights - Speaker View  
    C: Detailed Coverage List
    
    Returns the final DataFrame that was written to Excel.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from collections import Counter
    import string
    
    print("📊 Generating Executive Summary report...")
    
    # ==========================================
    # Section A: Overall Performance KPIs
    # ==========================================
    print("📈 Calculating Overall Performance KPIs...")
    
    # Get total call duration directly from the last row's segment end time
    if 'Segment End Time' in updated_call_df.columns and not updated_call_df.empty:
        # Get the last row's end time value directly (keep as string if it's string)
        total_duration = updated_call_df['Segment End Time'].iloc[-1]
        if pd.isna(total_duration):
            total_duration = "N/A"
        else:
            total_duration = str(total_duration)
    else:
        total_duration = "N/A"
    
    # Calculate coverage metrics
    total_points = len(results_df)
    covered_points = len(results_df[results_df['Covered'] == 'Covered'])
    uncovered_points = total_points - covered_points
    coverage_rate = (covered_points / total_points * 100) if total_points > 0 else 0
    
    kpis_data = {
        'Metric': [
            'Total Call Duration',
            'Compliance Coverage Rate (%)',
            'Total Points Checked',
            'Covered Points',
            'Uncovered Risk Points'
        ],
        'Value': [
            str(total_duration),
            f"{coverage_rate:.1f}%",
            str(total_points),
            str(covered_points),
            str(uncovered_points)
        ]
    }
    kpis_df = pd.DataFrame(kpis_data)
    
    # ==========================================
    # Section B: Key Insights - Speaker View
    # ==========================================
    print("🎤 Generating Speaker View insights...")
    
    # Filter out system recordings for speaker analysis
    non_system_df = updated_call_df[updated_call_df.get('System_Audio_Recording', 'NO') == 'NO'].copy()
    
    # Calculate speaker statistics
    speaker_stats = []
    if not non_system_df.empty and 'Speaker Roles' in non_system_df.columns and 'Transcription' in non_system_df.columns:
        # Group by speaker and calculate metrics
        for speaker in non_system_df['Speaker Roles'].unique():
            if pd.isna(speaker):
                continue
                
            speaker_data = non_system_df[non_system_df['Speaker Roles'] == speaker]
            
            # Calculate speaking duration using proper time parsing with strptime
            if 'Segment Start Time' in speaker_data.columns and 'Segment End Time' in speaker_data.columns:
                speaking_duration = 0.0
                
                for _, segment in speaker_data.iterrows():
                    start_time_str = str(segment.get('Segment Start Time', '')).strip()
                    end_time_str = str(segment.get('Segment End Time', '')).strip()
                    
                    if start_time_str and end_time_str and start_time_str != 'nan' and end_time_str != 'nan':
                        try:
                            from datetime import datetime
                            # Parse time strings as HH:MM:SS format
                            start_time = datetime.strptime(start_time_str, "%H:%M:%S")
                            end_time = datetime.strptime(end_time_str, "%H:%M:%S")
                            
                            # Calculate duration in seconds
                            duration_seconds = (end_time - start_time).total_seconds()
                            if duration_seconds >= 0:  # Only add positive durations
                                speaking_duration += duration_seconds
                                
                        except ValueError:
                            # If parsing fails, try to use the old parse_time_to_seconds as fallback
                            try:
                                start_seconds = parse_time_to_seconds(start_time_str)
                                end_seconds = parse_time_to_seconds(end_time_str)
                                duration_seconds = end_seconds - start_seconds
                                if duration_seconds >= 0:
                                    speaking_duration += duration_seconds
                            except:
                                continue  # Skip this segment if can't parse
                
                print(f"🔍 Debug - Speaker: {speaker}, Speaking Duration: {speaking_duration:.1f}s")
            else:
                speaking_duration = 0.0
                print(f"🔍 Debug - Speaker: {speaker} - Time columns not found")
            
            # Calculate word count and extract text
            all_text = ' '.join(speaker_data['Transcription'].fillna('').astype(str))
            word_count = len(all_text.replace(' ', ''))  # Character count for Chinese text
            
            # Calculate words per second
            words_per_second = (word_count / speaking_duration) if speaking_duration > 0 else 0
            
            # Generate top 5 keywords using pycantonese tokenization for Cantonese
            import pycantonese as pc
            # Tokenize using pycantonese and filter
            words = pc.segment(all_text)
            words = [w.strip() for w in words if len(w.strip()) >= 2]  # Keep words with 2+ characters
            
            # Use proper stopwords from dictionaries based on language
            stopwords = dictionaries.get_stopwords('CAN')  # Cantonese version
            filtered_words = [w for w in words if w not in stopwords]
            
            # Get top 10 most frequent words
            word_counter = Counter(filtered_words)
            top_10_words = [word for word, count in word_counter.most_common(10)]
            
            # Format as requested: "{'word1', 'word2', ..., 'word10'}"
            keywords_str = "{" + ", ".join([f"'{word}'" for word in top_10_words]) + "}"
            
            speaker_stats.append({
                'Speaker': speaker,
                'Speaking Duration (seconds)': speaking_duration,
                'Word Count': str(word_count),
                'Words per Second': f"{words_per_second:.2f}",
                'Top 10 Keywords': keywords_str
            })
    
    # Identify system recording speakers and add them to speaker_stats
    system_recording_speakers = set()
    if 'System_Audio_Recording' in updated_call_df.columns:
        system_df = updated_call_df[updated_call_df['System_Audio_Recording'] == 'YES']
        if not system_df.empty and 'Speaker Roles' in system_df.columns:
            system_recording_speakers = set(system_df['Speaker Roles'].dropna().unique())
            
            # Add system recording speakers to the speaker_stats
            for sys_speaker in system_recording_speakers:
                if sys_speaker and sys_speaker not in [stat['Speaker'] for stat in speaker_stats]:
                    speaker_stats.append({
                        'Role': 'System',
                        'Speaker': str(sys_speaker),
                        'Speaking Duration (seconds)': 'N/A',
                        'Word Count': 'N/A',
                        'Words per Second': 'N/A',
                        'Top 10 Keywords': 'N/A'
                    })
    
    # Sort speakers by word count to identify Sales and Customer (excluding system speakers)
    regular_speakers = [stat for stat in speaker_stats if stat.get('Role') != 'System']
    if regular_speakers:
        # Convert word count back to int for sorting
        for stat in regular_speakers:
            stat['_word_count_int'] = int(stat['Word Count'])
        
        regular_speakers.sort(key=lambda x: x['_word_count_int'], reverse=True)
        
        # Assign roles: highest word count = Sales, second highest = Customer
        if len(regular_speakers) >= 1:
            regular_speakers[0]['Role'] = 'Sales'
        if len(regular_speakers) >= 2:
            regular_speakers[1]['Role'] = 'Customer'
        
        # Remove the temporary sorting field and assign default roles
        for stat in regular_speakers:
            del stat['_word_count_int']
            if 'Role' not in stat:
                stat['Role'] = 'Other'
        
        # Ensure all speaker_stats have Role assigned
        for stat in speaker_stats:
            if 'Role' not in stat:
                stat['Role'] = 'Other'
    
    speaker_view_df = pd.DataFrame(speaker_stats)
    if not speaker_view_df.empty:
        # Reorder columns
        speaker_view_df = speaker_view_df[['Role', 'Speaker', 'Speaking Duration (seconds)', 'Word Count', 'Words per Second', 'Top 10 Keywords']]
    
    # ==========================================
    # Section C: Detailed Coverage List
    # ==========================================
    print("📋 Preparing Detailed Coverage List...")
    
    # Create a copy of results for modification
    detailed_coverage_df = results_df.copy()
    
    # Rename columns for better business understanding
    detailed_coverage_df = detailed_coverage_df.rename(columns={
        'Matched_Group': 'Matched_Group (Call Text)',
        'Best_Matching_Variation': 'Standard_Script'
    })
    
    # Enhance the Covered column with visual cues
    detailed_coverage_df['Covered'] = detailed_coverage_df['Covered'].apply(
        lambda x: f"✅ {x}" if x == 'Covered' else f"❌ {x}"
    )
    
    # Reorder columns to be more business-friendly
    business_friendly_columns = [
        'Required_Discussion_Point',
        'Standard_Script',
        'Covered',
        'Matched_Group (Call Text)',
        'Speaker',
        'Overlapping_Keywords',
        'Group_ID',
        'All_Variations_Count'
    ]
    
    # Add remaining columns (metric scores)
    remaining_columns = [col for col in detailed_coverage_df.columns if col not in business_friendly_columns]
    final_columns = business_friendly_columns + remaining_columns
    
    # Reorder the DataFrame
    detailed_coverage_df = detailed_coverage_df[final_columns]
    
    # Return the complete consolidated data for reference
    return {
        'kpis': kpis_df,
        'speaker_view': speaker_view_df,
        'detailed_coverage': detailed_coverage_df
    }

def write_executive_summary_content(ws, kpis_df, speaker_view_df, detailed_coverage_df):
    """Write Executive Summary content to a worksheet"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    import string
    
    current_row = 1
    
    # Define styles
    header_font = Font(bold=True, size=12)
    table_header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Section A: Overall Performance KPIs
    ws[f'A{current_row}'] = "A. Overall Performance KPIs"
    ws[f'A{current_row}'].font = header_font
    current_row += 2
    
    # Write KPIs data
    for idx, row in kpis_df.iterrows():
        ws[f'A{current_row}'] = row['Metric']
        ws[f'B{current_row}'] = row['Value']
        ws[f'A{current_row}'].font = table_header_font
        current_row += 1
    
    current_row += 2
    
    # Section B: Key Insights - Speaker View
    ws[f'A{current_row}'] = "B. Key Insights - Speaker View"
    ws[f'A{current_row}'].font = header_font
    current_row += 2
    
    if not speaker_view_df.empty:
        # Write headers
        for col_idx, col_name in enumerate(speaker_view_df.columns, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = col_name
            cell.font = table_header_font
            cell.fill = header_fill
            cell.border = border
        current_row += 1
        
        # Write data
        for idx, row in speaker_view_df.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.border = border
            current_row += 1
    else:
        ws[f'A{current_row}'] = "No speaker data available"
        current_row += 1
    
    current_row += 2
    
    # Section C: Detailed Coverage List
    ws[f'A{current_row}'] = "C. Detailed Coverage List"
    ws[f'A{current_row}'].font = header_font
    current_row += 2
    
    if not detailed_coverage_df.empty:
        # Write headers
        for col_idx, col_name in enumerate(detailed_coverage_df.columns, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = col_name
            cell.font = table_header_font
            cell.fill = header_fill
            cell.border = border
        
        detail_header_row = current_row
        current_row += 1
        
        # Write data
        for idx, row in detailed_coverage_df.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.border = border
            current_row += 1
        
        # Enable AutoFilter on the detailed coverage table
        detail_end_col = len(detailed_coverage_df.columns)
        detail_end_row = current_row - 1
        detail_end_col_letter = chr(ord('A') + detail_end_col - 1)
        ws.auto_filter.ref = f"A{detail_header_row}:{detail_end_col_letter}{detail_end_row}"
    else:
        ws[f'A{current_row}'] = "No coverage data available"
        current_row += 1
    
    # Adjust column widths for readability
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for very long content
        ws.column_dimensions[column_letter].width = adjusted_width

# ==========================================================
# Main Routine
# ----------------------------------------------------------
# Loads data, runs coverage check, and outputs results.
# ==========================================================
def run_analysis(call_file_path, script_file_path, script_sheet_name, output_folder, call_type=None, language=None, include_system_audio=False):
    """
    Core analysis function implementing the four-step strategy:
    1. Early Separation: Split call_df into analysis_df and system_audio_df
    2. Independent Processing: Only process analysis_df for grouping
    3. Focused Analysis: check_coverage only analyzes human dialogue
    4. Late Merge & Reporting: Combine results for final reports
    
    Args:
        call_file_path: Path to the call text file
        script_file_path: Path to the script file
        script_sheet_name: Name of the script sheet
        output_folder: Folder to save output files
        call_type: Type of call ('SQCCB' or 'Sales Call'), if None will auto-detect
        language: Language code ('CAN', 'ENG'), if None will auto-detect
        include_system_audio: Whether to include system audio recordings in analysis (default: False)
        
    Returns:
        dict: Results dictionary with status, output_file, and coverage_rate
    """
    # Start tracing memory allocations and overall timer
    tracemalloc.start()
    overall_start = print_checkpoint(1, "Program startup, loading data files")
    
    # 🔍 通话类型检测逻辑（单一事实来源）
    if call_type is None:
        # Fallback: 从文件名检测（仅在直接调用时使用）
        filename = os.path.basename(call_file_path).upper()
        if "SQCCB" in filename:
            call_type = "SQCCB"
        else:
            call_type = "Sales Call"
        print(f"⚠️  Fallback detection - 通话类型: {call_type}")
    else:
        print(f"✅ 接收到的通话类型: {call_type}")
    
    print(f"📁 文件名: {os.path.basename(call_file_path)}")
    
    # Check if files exist
    if not os.path.exists(call_file_path):
        print(f"❌ Error: Call text file '{call_file_path}' not found.")
        return {'status': 'ERROR', 'error': f'Call file not found: {call_file_path}', 'output_file': None, 'coverage_rate': 0.0}
        
    if not os.path.exists(script_file_path):
        print(f"❌ Error: Script file '{script_file_path}' not found.")
        return {'status': 'ERROR', 'error': f'Script file not found: {script_file_path}', 'output_file': None, 'coverage_rate': 0.0}
    
    try:
        step1_start = time.time()
        
        # Load call text file (CSV format)
        print(f"📁 Loading call text from: {call_file_path}")
        if call_file_path.endswith('.csv'):
            call_df = pd.read_csv(call_file_path)
            print(f"   Loading CSV file directly")
        elif call_file_path.endswith('.xlsx') or call_file_path.endswith('.xls'):
            xl_call = pd.ExcelFile(call_file_path, engine='openpyxl')
            call_df = xl_call.parse(xl_call.sheet_names[0])  # Use first sheet
            print(f"   Using sheet: {xl_call.sheet_names[0]}")
        else:
            raise ValueError(f"Unsupported file format. Please use .csv, .xlsx, or .xls files.")
        
        # Load script file
        if script_file_path.endswith('.xlsx'):
            xl_script = pd.ExcelFile(script_file_path, engine='openpyxl')
            script_df = xl_script.parse(script_sheet_name)
        else:
            script_df = pd.read_csv(script_file_path)

        # Apply optional composite discussion-point key for configured sheets
        script_df = apply_composite_point_key_if_configured(script_df, script_sheet_name)
            
        print_checkpoint(2, f"Excel文件加载完成 ({len(call_df):,} 通话行, {len(script_df)} 脚本行)", step1_start)
    except Exception as e:
        print(f"❌ 读取Excel文件错误: {e}")
        return {'status': 'ERROR', 'error': str(e), 'output_file': None, 'coverage_rate': 0.0}

    # Initialize the checker
    step2_start = time.time()
    checker = AdvancedCallCoverageChecker()
    
    # 🔍 语言检测逻辑（单一事实来源）
    if language is None:
        # Fallback: 从文件名检测语言（仅在直接调用时使用）
        filename = os.path.basename(call_file_path).upper()
        if "_C" in filename:
            language = "CAN"
        elif "_E" in filename:
            language = "ENG"
        else:
            language = "CAN"  # 默认粤语
        print(f"⚠️  Fallback detection - 语言: {language}")
    else:
        print(f"✅ 接收到的语言: {language}")
    
    # 设置checker的语言
    checker.current_language = language
    
    # Load call-specific weights for all call types (unified processing)
    # This function is intelligent enough to handle both Sales Call and SQCCB based on script_sheet_name
    checker.load_call_specific_weights(call_file_path, script_df, script_sheet_name)
    
    # Check if the detected language is supported
    if checker.current_language not in ['CAN', 'ENG']:
        print(f"❌ Error: Language '{checker.current_language}' is not supported in Cantonese version.")
        print(f"Cantonese version only supports CAN (Cantonese) and ENG (English).")
        return {'status': 'ERROR', 'error': f'Unsupported language: {checker.current_language}', 'output_file': None, 'coverage_rate': 0.0}
    
    print(f"✅ Language validation passed: {checker.current_language}")
    print_checkpoint(3, "初始化分析器并验证语言支持", step2_start)
    
    # STEP 1 & 2: UNIFIED PROCESSING - Build grouped lines with call_type-aware system recording detection
    step3_start = time.time()
    print(f"\n📍 统一处理 - 使用 {call_type} 模式进行分组...")
    
    # Build grouped lines using the unified function that handles system recording detection internally
    grouped_lines, original_to_group, updated_call_df, system_recording_flags = checker.build_grouped_lines(call_df, call_type=call_type, include_system_audio=include_system_audio)
    
    # Create analysis_df and system_audio_df from updated_call_df (no duplicate calculation)
    analysis_df = updated_call_df[updated_call_df['System_Audio_Recording'] == 'NO'].copy()
    system_audio_df = updated_call_df[updated_call_df['System_Audio_Recording'] == 'YES'].copy()
    
    print(f"✅ 分组处理完成:")
    print(f"   - 创建了 {len(grouped_lines)} 个组")
    print_checkpoint(4, f"创建并保存分组通话数据 ({len(grouped_lines)} 个分组)", step3_start)
    
    # STEP 3: FOCUSED ANALYSIS - Coverage analysis on human dialogue only
    step5_start = time.time()
    print("\n📍 步骤3: 专注分析 - 对人类对话进行覆盖率分析...")
    
    # Check coverage using human-dialogue corpus + human-dialogue groups
    results = checker.check_coverage(analysis_df, script_df, grouped_lines, threshold=0.3)
    
    print(f"✅ 覆盖率分析完成:")
    print(f"   - 分析了 {len(results)} 个讨论点")
    print(f"   - 使用了 {len(grouped_lines)} 个人类对话组")
    print_checkpoint(5, f"完成覆盖分析 ({len(results)} 个要点)", step5_start)
    
    # STEP 4: LATE MERGE & REPORTING - Combine results for final reports
    step6_start = time.time()
    print("\n📍 步骤4: 后期合并与报告 - 创建综合报告...")
    
    # Create call text analysis with proper merging
    call_text_analysis = checker.create_call_text_analysis_view_with_separation(
        grouped_lines, system_audio_df, script_df, threshold=0.3
    )
    
    # Create sentence level output
    sentence_level_output = checker.create_sentence_level_output(updated_call_df, original_to_group)
    print_checkpoint(6, f"生成最终分析视图", step6_start)
    
    # Generate Reports with Executive Summary
    step7_start = time.time()
    print("\n📊 生成包含Executive Summary的最终报告...")
    
    # Generate unique output filename based on input filename and save to output_folder
    input_filename = os.path.basename(call_file_path)
    if input_filename.endswith('.csv'):
        # Remove .wav if present, then replace .csv with .xlsx
        if '.wav' in input_filename:
            output_filename = input_filename.replace('.wav.csv', '.xlsx')
        else:
            output_filename = input_filename.replace('.csv', '.xlsx')
    elif input_filename.endswith('.xlsx') or input_filename.endswith('.xls'):
        # Remove .wav if present, then add _result
        if '.wav' in input_filename:
            output_filename = input_filename.replace('.wav.xlsx', '.xlsx').replace('.wav.xls', '.xlsx')
        else:
            output_filename = input_filename.replace('.xlsx', '_result.xlsx').replace('.xls', '_result.xlsx')
    else:
        # Remove .wav if present, then add _result
        if '.wav' in input_filename:
            output_filename = input_filename.replace('.wav', '') + '_result.xlsx'
        else:
            output_filename = f"{input_filename}_result.xlsx"
    
    # Create full output path
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, output_filename)
    
    print(f"📝 Output file will be: {output_file}")
    try:
        # Generate the Executive Summary data first
        summary_data = generate_executive_summary_report(
            results, updated_call_df, script_df, grouped_lines, None  # Don't save to separate file
        )
        
        # Save to the main output file with Executive Summary as first sheet
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # First, create Executive Summary sheet manually
            wb = writer.book
            ws = wb.create_sheet("Executive Summary", 0)
            write_executive_summary_content(ws, summary_data['kpis'], summary_data['speaker_view'], summary_data['detailed_coverage'])
            
            # Then write other sheets
            call_text_analysis.to_excel(writer, sheet_name='Coverage Analysis', index=False)
            sentence_level_output.to_excel(writer, sheet_name='Original Call Text', index=False)
            
            # Remove the default 'Sheet' if it exists
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
        
        # Verify file creation
        if os.path.exists(output_file):
            file_size = os.path.getsize(output_file)
            print(f"✅ Excel文件成功创建: {output_file} ({file_size:,} bytes)")
        else:
            print(f"❌ Excel文件未创建: {output_file}")
            return {'status': 'ERROR', 'error': 'File creation failed', 'output_file': None, 'coverage_rate': 0.0}
            
    except Exception as e:
        print(f"❌ 生成报告错误: {e}")
        print("尝试保存备份CSV文件...")
        backup_folder = os.path.join(output_folder, 'backup')
        os.makedirs(backup_folder, exist_ok=True)
        results.to_csv(os.path.join(backup_folder, f'coverage_analysis_backup_{os.path.splitext(output_filename)[0]}.csv'), index=False, encoding='utf-8-sig')
        call_text_analysis.to_csv(os.path.join(backup_folder, f'call_text_analysis_backup_{os.path.splitext(output_filename)[0]}.csv'), index=False, encoding='utf-8-sig')
        sentence_level_output.to_csv(os.path.join(backup_folder, f'sentence_level_analysis_backup_{os.path.splitext(output_filename)[0]}.csv'), index=False, encoding='utf-8-sig')
        return {'status': 'ERROR', 'error': str(e), 'output_file': None, 'coverage_rate': 0.0}
    
    # Create grouped call DataFrame for reference
    grouped_call_df = pd.DataFrame(grouped_lines)
    reference_file = os.path.join(output_folder, f"grouped_call_data_{os.path.splitext(output_filename)[0]}.xlsx")
    grouped_call_df.to_excel(reference_file, index=False)
    
    print_checkpoint(7, f"生成包含执行摘要的报告 ({output_file})", step7_start)
    
    # Calculate final metrics
    total_time = time.time() - overall_start
    covered_count = len(results[results['Covered'] == 'Covered'])
    total_points = len(results)
    coverage_rate = (covered_count / total_points * 100) if total_points > 0 else 0.0
    
    print(f"\n🎉 四步早期分离策略分析完成! 总时间: {total_time:.1f}s")
    print(f"   • STEP 1: Early Separation - System recording detection")
    print(f"   • STEP 2: Independent Processing - Human dialogue grouping")
    print(f"   • STEP 3: Focused Analysis - Coverage analysis on human dialogue only")
    print(f"   • STEP 4: Late Merge & Reporting - Comprehensive report generation")
    print(f"   📋 Files created:")
    print(f"   • {output_file}: Executive Summary with late merge strategy")
    print(f"   • {reference_file}: Pure human dialogue groups")
    
    print(f"\n📊 Coverage Summary:")
    print(f"   • Coverage Rate: {coverage_rate:.1f}% ({covered_count}/{total_points} points)")
    print(f"   • Total groups created: {len(grouped_lines)}")
    print(f"   • Call type: {call_type}")

    print(f"\n📊 Output Files Created:")
    print(f"  • {output_file} (3 sheets with Executive Summary):")
    print(f"    - Executive Summary: Consolidated business-friendly report with KPIs, Speaker View & Detailed Coverage")
    print(f"    - Coverage Analysis: {len(call_text_analysis)} call text segments analysis")
    print(f"    - Original Call Text: {len(sentence_level_output)} sentence-level analysis")
    print(f"  • {reference_file}: {len(grouped_call_df)} grouped call data reference")
    print(f"\n🎉 Executive Summary Features:")
    print(f"  • 📈 Overall Performance KPIs: Total duration, coverage rate, risk points")
    print(f"  • 🎤 Speaker View: Role identification (Sales/Customer/System), duration, keywords")
    print(f"  • 📋 Detailed Coverage List: Business-friendly format with ✅/❌ indicators")
    print(f"  • 💼 Professional Excel formatting: Bold headers, borders, auto-filter")
    print(f"  • 🔄 Dynamic filename: {input_filename} → {output_filename}")

    # Final checkpoint and cleanup
    overall_time = print_checkpoint("FINAL", "All processing completed", overall_start)
    
    # Print memory usage
    current, peak = tracemalloc.get_traced_memory()
    print(f"\n💾 Memory usage summary:")
    print(f"  - Current memory usage: {current / 10**6:.2f}MB")
    print(f"  - Peak memory usage: {peak / 10**6:.2f}MB")

    # Stop tracing memory allocations
    tracemalloc.stop()
    
    print(f"\n🎉 Program execution completed! Total time: {overall_time - overall_start:.2f}s")
    
    # Extract speaker word counts from summary_data
    sales_word_count = 0
    customer_word_count = 0
    
    if summary_data and 'speaker_view' in summary_data:
        speaker_view_df = summary_data['speaker_view']
        if not speaker_view_df.empty:
            for _, row in speaker_view_df.iterrows():
                if row.get('Role') == 'Sales' and pd.notna(row.get('Word Count')) and str(row.get('Word Count')).isdigit():
                    sales_word_count = int(row['Word Count'])
                elif row.get('Role') == 'Customer' and pd.notna(row.get('Word Count')) and str(row.get('Word Count')).isdigit():
                    customer_word_count = int(row['Word Count'])
    
    # Convert total_call_duration to seconds (numeric)
    total_call_duration_seconds = 0
    if summary_data and 'kpis' in summary_data:
        kpis_df = summary_data['kpis']
        duration_row = kpis_df[kpis_df['Metric'] == 'Total Call Duration']
        if not duration_row.empty:
            duration_value = duration_row.iloc[0]['Value']
            if pd.notna(duration_value) and str(duration_value) != 'N/A':
                try:
                    # Parse time string to seconds
                    duration_str = str(duration_value)
                    if ':' in duration_str:
                        parts = duration_str.split(':')
                        if len(parts) == 2:  # MM:SS
                            total_call_duration_seconds = int(parts[0]) * 60 + int(parts[1])
                        elif len(parts) == 3:  # HH:MM:SS
                            total_call_duration_seconds = int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
                    else:
                        # Direct numeric value
                        total_call_duration_seconds = float(duration_str)
                except (ValueError, TypeError):
                    total_call_duration_seconds = 0
    
    # Return enhanced results dictionary
    return {
        'status': 'SUCCESS',
        'output_file': output_file,
        'coverage_rate': coverage_rate,
        'total_points': total_points,
        'covered_points': covered_count,
        'total_call_duration': total_call_duration_seconds,
        'sales_word_count': sales_word_count,
        'customer_word_count': customer_word_count,
        'processing_time': total_time,
        'call_type': call_type,
        'language': checker.current_language
    }


def main():
    """
    Main function for standalone testing using default configuration.
    Note: When used standalone, you need to modify the paths below or use run_batch_analysis.py
    """
    # Default paths for standalone testing - modify these as needed
    default_call_file = "call_text_sample_C.wav.csv"
    default_script_file = "Scripts.xlsx"
    default_script_sheet = "Script"
    default_output_folder = "."
    
    print("⚠️  Running in standalone mode with default paths.")
    print("For batch processing, use run_batch_analysis.py instead.")
    
    return run_analysis(default_call_file, default_script_file, default_script_sheet, default_output_folder)


# Standard Python entry point
if __name__ == "__main__":
    main()
